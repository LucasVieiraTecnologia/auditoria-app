# ============================================================
# AUDITORIA INTELIGENTE DE CONDOMÍNIO
# ============================================================

import os
import re
import fitz
import hashlib
import time
import pandas as pd
import pytesseract
import plotly.express as px
import plotly.graph_objects as go
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import unquote_plus, urljoin, urlparse
from urllib.request import Request, urlopen
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from dotenv import load_dotenv
load_dotenv()
from openpyxl.drawing.image import Image as XLImage

# ============================================================
# CONFIGURAÇÕES
# ============================================================

PASTA_PDFS = "./ArquivosPDF"
PASTA_IMAGENS = "./notas_fiscais"
PASTA_DOCUMENTOS_NF = "./documentos_nf"

ARQUIVO_EXCEL = "auditoria_condominio.xlsx"

ARQUIVO_CSV_BALANCO = "auditoria_balanco_mensal.csv"
ARQUIVO_CSV_CATEGORIAS = "auditoria_categorias_mensais.csv"
ARQUIVO_CSV_MOVIMENTACOES = "auditoria_movimentacoes.csv"
ARQUIVO_CSV_COBRANCAS = "auditoria_composicao_cobrancas.csv"

os.makedirs(PASTA_IMAGENS, exist_ok=True)
os.makedirs(PASTA_PDFS, exist_ok=True)
os.makedirs(PASTA_DOCUMENTOS_NF, exist_ok=True)
NF_DOWNLOAD_DELAY = float(os.getenv("NF_DOWNLOAD_DELAY", "1.2"))
NF_DOWNLOAD_TIMEOUT = int(os.getenv("NF_DOWNLOAD_TIMEOUT", "25"))
NF_MAX_DOWNLOADS = int(os.getenv("NF_MAX_DOWNLOADS", "300"))
NF_MAX_BYTES = int(os.getenv("NF_MAX_BYTES", str(25 * 1024 * 1024)))
_NF_DOWNLOAD_CACHE = {}
_NF_DOWNLOAD_COUNT = 0
_NF_LAST_DOWNLOAD = 0.0

# ============================================================
# REGEX
# ============================================================

REGEX_DATA = re.compile(r"\d{2}/\d{2}/\d{4}")
REGEX_VALOR_TXT = r"\(?-?\d{1,3}(?:\.\d{3})*,\d{2}\)?"
REGEX_VALOR = re.compile(REGEX_VALOR_TXT)
REGEX_CNPJ = re.compile(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}")
REGEX_URL = re.compile(r'https?://[^\s\])}>"]+', re.IGNORECASE)
REGEX_CHAVE_NFE = re.compile(r"(?:\d[\s\.\-]?){44}")
REGEX_NUMERO_NF = re.compile(r"(?:NF(?:-?e|S-?e)?|Nota\s+Fiscal|DANFE|N[úu]mero\s+da\s+Nota|N[ºo]\.?\s+da\s+Nota)\D{0,35}([0-9]{3,12})", re.IGNORECASE)
REGEX_PERIODO = re.compile(
    r"(?:Período|Periodo|De)\s+(\d{2}/\d{2}/\d{4})\s+(?:a|até|ate)\s+(\d{2}/\d{2}/\d{4})",
    re.IGNORECASE
)

# ============================================================
# COLUNAS PADRÃO
# ============================================================

COLUNAS_BALANCO = [
    "Arquivo_Origem", "Condominio", "Periodo_Inicio", "Periodo_Fim", "Ano",
    "Mes", "Mes_Ano", "Saldo_Anterior", "Total_Receitas", "Total_Despesas",
    "Movimento_Liquido", "Saldo_Final", "Diferenca_Resultado", "Diferenca_Saldo",
    "Status_Validacao", "Hash",
]

COLUNAS_CATEGORIAS = [
    "Arquivo_Origem", "Condominio", "Periodo_Inicio", "Periodo_Fim", "Ano",
    "Mes", "Mes_Ano", "Grupo", "Categoria", "Valor", "Valor_Entrada",
    "Valor_Saida", "Valor_Resultante", "Origem_Extracao", "Hash",
]

COLUNAS_MOVIMENTACOES = [
    "Arquivo_Origem", "Condominio", "Periodo_Inicio", "Periodo_Fim", "Ano",
    "Mes", "Mes_Ano", "Data", "Conta", "Tipo_Movimento", "Categoria",
    "Fornecedor", "Descricao", "Valor_Real", "Valor_Assinado", "Valor_Entrada",
    "Valor_Saida", "Valor_Transferencia", "Saldo_Apos", "CNPJ", "PIX",
    "Score_Suspeita", "Motivo_Suspeita", "Numero_NF", "Chave_NF",
    "Status_Consulta_NF", "Link_Consulta_NF", "Link_Origem_NF",
    "Arquivo_NF_Baixado", "Link_NF", "Origem_Extracao", "Hash",
]

COLUNAS_COBRANCAS = [
    "Arquivo_Origem", "Condominio", "Periodo_Inicio", "Periodo_Fim", "Ano",
    "Mes", "Mes_Ano", "Unidade", "Bloco", "Codigo_Cobranca", "Descricao",
    "Categoria", "Vencimento", "Data_Credito", "Data_Liquidacao", "Valor",
    "Valor_Assinado", "Tipo_Movimento", "Origem_Extracao", "Hash",
]

COLUNAS_VALIDACAO = [
    "Arquivo_Origem", "Mes_Ano", "Receitas_Balancete", "Entradas_Movimentacao",
    "Diferenca_Entradas", "Despesas_Balancete", "Saidas_Movimentacao",
    "Diferenca_Saidas", "Movimento_Liquido_Balancete", "Saldo_Anterior",
    "Saldo_Final", "Status",
]

# ============================================================
# UTILITÁRIOS
# ============================================================

def limpar_texto(texto):
    if not texto: return ""
    texto = str(texto).replace("\n", " ").replace("\r", " ")
    texto = re.sub(r"\s+", " ", texto)
    return re.sub(r"_+", "", texto).strip()

def limpar_linha(texto):
    if not texto: return ""
    texto = str(texto).replace("\r", " ")
    return re.sub(r"\s+", " ", texto).strip()

def normalizar(texto):
    if texto is None: return ""
    texto = str(texto).lower()
    troca = {"á": "a", "à": "a", "ã": "a", "â": "a", "é": "e", "ê": "e", "í": "i", "ó": "o", "ô": "o", "õ": "o", "ú": "u", "ü": "u", "ç": "c"}
    for original, novo in troca.items(): texto = texto.replace(original, novo)
    return texto

def valor_para_float(valor):
    if valor is None: return None
    valor = str(valor).strip()
    negativo = "(" in valor and ")" in valor or valor.startswith("-")
    valor = valor.replace("(", "").replace(")", "").replace("-", "").replace(".", "").replace(",", ".")
    try:
        valor_float = float(valor)
    except Exception: return None
    if negativo: valor_float *= -1
    return valor_float

def float_para_br(valor):
    try:
        if pd.isna(valor): return ""
        return f"{float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception: return ""

def converter_data(data_txt):
    try: return pd.to_datetime(data_txt, format="%d/%m/%Y", errors="coerce")
    except Exception: return pd.NaT

def gerar_hash(texto):
    return hashlib.md5(str(texto).encode("utf-8")).hexdigest()

def nome_arquivo_seguro(texto, limite=90):
    texto = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(texto or "")).strip("_")
    return texto[:limite] or "arquivo"

def criar_dataframe(registros, colunas):
    if registros:
        df = pd.DataFrame(registros)
        for col in colunas:
            if col not in df.columns: df[col] = None
        return df[colunas]
    return pd.DataFrame(columns=colunas)

# ============================================================
# FUNÇÕES CRÍTICAS PARA NÃO PEGAR PERCENTUAL COMO VALOR
# ============================================================

def extrair_valores_com_contexto(texto):
    resultados = []
    for m in REGEX_VALOR.finditer(texto):
        valor_txt = m.group(0)
        valor_float = valor_para_float(valor_txt)
        if valor_float is None: continue
        pos_fim = m.end()
        restante = texto[pos_fim:pos_fim + 5]
        eh_percentual = "%" in restante
        resultados.append({"txt": valor_txt, "valor": valor_float, "eh_percentual": eh_percentual})
    return resultados

def escolher_valor_monetario(texto):
    valores = extrair_valores_com_contexto(texto)
    if not valores: return None
    nao_percentuais = [item["valor"] for item in valores if not item["eh_percentual"]]
    if not nao_percentuais: nao_percentuais = [item["valor"] for item in valores]
    if not nao_percentuais: return None
    return max(nao_percentuais, key=lambda x: abs(x))

def extrair_trecho_entre_rotulos(texto, rotulo_inicio, rotulos_fim):
    texto_norm = normalizar(texto)
    inicio_norm = normalizar(rotulo_inicio)
    pos_ini = texto_norm.find(inicio_norm)
    if pos_ini == -1: return ""
    pos_busca = pos_ini + len(inicio_norm)
    pos_fim_encontrado = None
    for rotulo_fim in rotulos_fim:
        fim_norm = normalizar(rotulo_fim)
        pos_fim = texto_norm.find(fim_norm, pos_busca)
        if pos_fim != -1:
            if pos_fim_encontrado is None or pos_fim < pos_fim_encontrado:
                pos_fim_encontrado = pos_fim
    if pos_fim_encontrado is None: return texto[pos_ini:]
    return texto[pos_ini:pos_fim_encontrado]

def buscar_valor_rotulo_por_trecho(texto, rotulo, proximos_rotulos):
    trecho = extrair_trecho_entre_rotulos(texto, rotulo, proximos_rotulos)
    return escolher_valor_monetario(trecho) if trecho else None

# ============================================================
# OCR E EXTRAÇÃO
# ============================================================

def executar_ocr(pagina):
    try:
        pix = pagina.get_pixmap(matrix=fitz.Matrix(2, 2))
        temp = "ocr_temp.png"
        pix.save(temp, output="png")
        texto = pytesseract.image_to_string(Image.open(temp), lang="por")
        os.remove(temp)
        return texto
    except Exception: return ""

def extrair_paginas_pdf(pdf_path):
    doc = fitz.open(pdf_path)
    paginas = []
    for idx in range(len(doc)):
        pagina = doc[idx]
        texto = pagina.get_text("text")
        if len(texto.strip()) < 30: texto = executar_ocr(pagina)
        links = []
        for link in pagina.get_links():
            uri = link.get("uri")
            if uri and uri not in links:
                links.append(uri)
        for uri in REGEX_URL.findall(texto):
            uri = uri.rstrip('.,;)]}')
            if uri and uri not in links:
                links.append(uri)
        linhas = [limpar_linha(linha) for linha in texto.splitlines() if limpar_linha(linha)]
        paginas.append({
            "pagina": idx + 1, "texto": texto, "texto_limpo": limpar_texto(texto), "linhas": linhas, "links": links
        })
    return paginas

def extrair_metadados(pdf_path, paginas):
    texto_total = "\n".join(pagina["texto"] for pagina in paginas[:10])
    texto_limpo = limpar_texto(texto_total)
    periodo_ini, periodo_fim, ano, mes, mes_ano = None, None, None, None, None
    m_periodo = REGEX_PERIODO.search(texto_limpo)
    if m_periodo:
        periodo_ini = m_periodo.group(1)
        periodo_fim = m_periodo.group(2)
        dt = converter_data(periodo_fim)
        if pd.notna(dt):
            ano, mes, mes_ano = int(dt.year), int(dt.month), dt.strftime("%Y-%m")

    condominio = "NÃO IDENTIFICADO"
    m_condominio = re.search(r"Condomínio:\s*([A-ZÁÉÍÓÚÂÊÔÃÕÇ0-9\s]+)", texto_limpo, re.IGNORECASE)
    if m_condominio: condominio = limpar_texto(m_condominio.group(1))
    if condominio == "NÃO IDENTIFICADO":
        m_condominio_alt = re.search(r"CONDOMÍNIO\s+RESIDENCIAL\s+([A-ZÁÉÍÓÚÂÊÔÃÕÇ0-9\s]+)", texto_limpo, re.IGNORECASE)
        if m_condominio_alt: condominio = "CONDOMÍNIO RESIDENCIAL " + limpar_texto(m_condominio_alt.group(1))

    return {
        "Arquivo_Origem": pdf_path.name, "Condominio": condominio, "Periodo_Inicio": periodo_ini,
        "Periodo_Fim": periodo_fim, "Ano": ano, "Mes": mes, "Mes_Ano": mes_ano
    }

# ============================================================
# BALANÇO E CATEGORIAS
# ============================================================

def buscar_saldo_anterior(texto):
    trecho = extrair_trecho_entre_rotulos(texto, "Saldo Anterior", ["Receitas", "Total de Receitas", "Despesas Mensais"])
    return escolher_valor_monetario(trecho) if trecho else None

def buscar_total_receitas(texto):
    trecho = extrair_trecho_entre_rotulos(texto, "Total de Receitas", ["Despesas Mensais", "Despesas", "Total de Mensais", "Resumo Financeiro", "SALDO ATUAL"])
    return abs(escolher_valor_monetario(trecho)) if trecho else None

def buscar_total_despesas(texto):
    trecho = extrair_trecho_entre_rotulos(texto, "Total de Despesas", ["Mov. Líquido", "Mov. Liquido", "Saldo em", "SALDO ATUAL", "Resumo Financeiro", "AHAVA"])
    return abs(escolher_valor_monetario(trecho)) if trecho else None

def buscar_movimento_liquido(texto):
    trecho = extrair_trecho_entre_rotulos(texto, "Mov. Líquido", ["Saldo em", "SALDO ATUAL", "Resumo Financeiro", "AHAVA"])
    if not trecho: trecho = extrair_trecho_entre_rotulos(texto, "Mov. Liquido", ["Saldo em", "SALDO ATUAL", "Resumo Financeiro", "AHAVA"])
    return escolher_valor_monetario(trecho) if trecho else None

def buscar_saldo_final(texto, meta):
    periodo_fim = meta.get("Periodo_Fim")
    rotulos = [f"Saldo em {periodo_fim}"] if periodo_fim else []
    rotulos.extend(["SALDO ATUAL (SALDO ANTERIOR + RECEITA - DESPESA)", "Saldo Atual", "Saldo Total"])
    for rotulo in rotulos:
        trecho = extrair_trecho_entre_rotulos(texto, rotulo, ["Resumo dos Saldos", "AHAVA", "Conta Saldo", "Banco Sicoob", "Emitido em"])
        if trecho:
            valor = escolher_valor_monetario(trecho)
            if valor is not None: return valor
    return None

def extrair_balanco_mensal(pdf_path, paginas, meta):
    texto_primeiras = "\n".join(pagina["texto_limpo"] for pagina in paginas[:8])
    saldo_anterior = buscar_saldo_anterior(texto_primeiras)
    total_receitas = buscar_total_receitas(texto_primeiras)
    total_despesas = buscar_total_despesas(texto_primeiras)
    movimento_liquido = buscar_movimento_liquido(texto_primeiras)
    saldo_final = buscar_saldo_final(texto_primeiras, meta)

    if movimento_liquido is None and total_receitas is not None and total_despesas is not None:
        movimento_liquido = round(total_receitas - total_despesas, 2)
    if saldo_final is None and saldo_anterior is not None and movimento_liquido is not None:
        saldo_final = round(saldo_anterior + movimento_liquido, 2)

    diferenca_resultado, diferenca_saldo = None, None
    if total_receitas is not None and total_despesas is not None and movimento_liquido is not None:
        diferenca_resultado = round(movimento_liquido - (total_receitas - total_despesas), 2)
    if saldo_anterior is not None and movimento_liquido is not None and saldo_final is not None:
        diferenca_saldo = round(saldo_final - (saldo_anterior + movimento_liquido), 2)

    status_validacao = "OK"
    if diferenca_resultado not in [None, 0]: status_validacao = "DIVERGÊNCIA_RESULTADO"
    if diferenca_saldo not in [None, 0]: status_validacao = "DIVERGÊNCIA_SALDO"

    return {
        **meta, "Saldo_Anterior": saldo_anterior, "Total_Receitas": total_receitas,
        "Total_Despesas": total_despesas, "Movimento_Liquido": movimento_liquido,
        "Saldo_Final": saldo_final, "Diferenca_Resultado": diferenca_resultado,
        "Diferenca_Saldo": diferenca_saldo, "Status_Validacao": status_validacao,
        "Hash": gerar_hash(f"{pdf_path.name}_{meta.get('Mes_Ano')}_{saldo_anterior}_{total_receitas}_{total_despesas}_{saldo_final}")
    }

RECEITAS_ROTULOS = ["Cotas do Mês", "Juros", "Multas", "Tarifa bancária", "Água", "Atualização Monetária", "Honorário de Cobrança", "Pagamentos a Maior", "Fundo de Obras", "Taxa fixa Copasa", "Outras Entradas"]
DESPESAS_ROTULOS = ["Energia elétrica", "Água e Esgoto", "Seguro obrigatório", "Limpeza e conservação", "Limpeza Caixas de gordura - Água", "Câmeras e Interfones", "Manutenção Geral", "Locação de Equipamentos", "Dedetização - Controle de Pragas", "Combustível", "Conservadora", "Administradora de Condomínio", "Síndico Profissional", "Assessoria e Serviços Profissionais", "Chaveiro", "Taxa de Administração de Água e-ou Gás", "INSS Notas", "Honorários de cobrança", "Despesa com Processo Judicial", "Tarifa bancária", "Reformas em geral"]

def extrair_secao(texto, inicio, fim=None):
    texto_norm, inicio_norm = normalizar(texto), normalizar(inicio)
    pos_ini = texto_norm.find(inicio_norm)
    if pos_ini == -1: return ""
    if fim:
        pos_fim = texto_norm.find(normalizar(fim), pos_ini + len(inicio_norm))
        if pos_fim != -1: return texto[pos_ini:pos_fim]
    return texto[pos_ini:]

def extrair_categorias_mensais(pdf_path, paginas, meta):
    registros = []
    texto = "\n".join(pagina["texto_limpo"] for pagina in paginas[:8])
    secao_receitas = extrair_secao(texto, "Receitas", "Despesas Mensais")
    secao_despesas = extrair_secao(texto, "Despesas Mensais", "Resumo Financeiro")
    if not secao_despesas: secao_despesas = extrair_secao(texto, "Despesas Mensais")

    for idx, rotulo in enumerate(RECEITAS_ROTULOS):
        proximos = RECEITAS_ROTULOS[idx + 1:] + ["Total de Receitas", "Despesas Mensais"]
        valor = buscar_valor_rotulo_por_trecho(secao_receitas, rotulo, proximos)
        if valor is not None:
            registros.append({**meta, "Grupo": "Entrada", "Categoria": rotulo, "Valor": abs(valor), "Valor_Entrada": abs(valor), "Valor_Saida": 0, "Valor_Resultante": valor, "Origem_Extracao": "Balancete/Categorias", "Hash": gerar_hash(f"{pdf_path.name}_{meta.get('Mes_Ano')}_Entrada_{rotulo}_{valor}")})

    for idx, rotulo in enumerate(DESPESAS_ROTULOS):
        proximos = DESPESAS_ROTULOS[idx + 1:] + ["Total de Mensais", "Total de Manutenção", "Total de Despesas Gerais", "Total de Prestação de serviços", "Total de Impostos e Taxas", "Total de Jurídico ou Cobrança", "Total de Despesas Bancárias", "Total de Despesas Obras e melhorias", "Total de Despesas", "Mov. Líquido", "Mov. Liquido"]
        valor = buscar_valor_rotulo_por_trecho(secao_despesas, rotulo, proximos)
        if valor is not None:
            registros.append({**meta, "Grupo": "Saída", "Categoria": rotulo, "Valor": abs(valor), "Valor_Entrada": 0, "Valor_Saida": abs(valor), "Valor_Resultante": -abs(valor), "Origem_Extracao": "Balancete/Categorias", "Hash": gerar_hash(f"{pdf_path.name}_{meta.get('Mes_Ano')}_Saida_{rotulo}_{valor}")})

    return registros

# ============================================================
# CLASSIFICAÇÃO E MOVIMENTAÇÃO
# ============================================================

def classificar_categoria(descricao):
    t = normalizar(descricao)
    if "cemig" in t or "energia" in t: return "Energia"
    if "copasa" in t or "agua" in t or "esgoto" in t or "taxa fixa copasa" in t: return "Água"
    if "cota" in t: return "Cotas do Mês"
    if "fundo de obras" in t: return "Fundo de Obras"
    if "fundo de reserva" in t: return "Fundo de Reserva"
    if "tarifa bancaria" in t: return "Tarifa bancária"
    if "juros" in t: return "Juros"
    if "multa" in t: return "Multas"
    if "honorario" in t or "cobranca" in t: return "Honorários/Cobrança"
    if "receita federal" in t or "inss" in t or "darf" in t or "ministerio da fazenda" in t: return "Impostos"
    if any(k in t for k in ["manutencao", "reparo", "limpeza", "material", "dedetizacao", "interfone", "fechadura", "obra", "reforma", "cacamba", "locacamba"]): return "Manutenção/Obras"
    if "seguro" in t or "tokio" in t or "allianz" in t: return "Seguro"
    if "sindico" in t or "administradora" in t or "ahava" in t or "conservadora" in t: return "Prestação de Serviços"
    if "tribunal" in t or "processo judicial" in t: return "Jurídico"
    if "transferencia" in t or "transf." in t: return "Transferência"
    return "Outros"

def extrair_fornecedor(descricao):
    texto = descricao.upper()
    if " - " in descricao: return descricao.split(" - ")[0].strip()[:100]
    fornecedores = ["CEMIG", "COPASA", "RECEITA FEDERAL", "TOKIO MARINE", "ALLIANZ", "CONSERVADORA ATUANTE", "AHAVA", "TODAH", "MATOSO", "MGPRAG", "LOCAÇAMBA", "LOCACAMBA", "TRIBUNAL DE JUSTIÇA", "TRIBUNAL DE JUSTICA", "CHARLES", "PAULO HENRIQUE", "JOSE LUCIO", "VINICIUS SANTANA", "DELMA"]
    for f in fornecedores:
        if f in texto: return f
    return ""

def extrair_cnpj(texto):
    m = REGEX_CNPJ.search(texto)
    return m.group() if m else ""

def detectar_pix(texto):
    t = normalizar(texto)
    return any(termo in t for termo in ["pix", "chave pix", "transferencia", "ted", "cpf"])

def somente_digitos(texto):
    return re.sub(r"\D", "", str(texto or ""))

def validar_chave_nfe(chave):
    chave = somente_digitos(chave)
    if len(chave) != 44 or len(set(chave)) == 1:
        return False
    corpo, dv_informado = chave[:43], int(chave[43])
    peso, soma = 2, 0
    for digito in reversed(corpo):
        soma += int(digito) * peso
        peso = 2 if peso == 9 else peso + 1
    resto = soma % 11
    dv_calculado = 0 if resto in (0, 1) else 11 - resto
    return dv_calculado == dv_informado

def extrair_chaves_nfe(texto):
    chaves = []
    for m in REGEX_CHAVE_NFE.finditer(str(texto or "")):
        chave = somente_digitos(m.group(0))
        if len(chave) == 44 and chave not in chaves:
            chaves.append(chave)
    return chaves

def numero_nf_da_chave(chave):
    chave = somente_digitos(chave)
    if len(chave) != 44:
        return ""
    numero = chave[25:34].lstrip("0")
    return numero or chave[25:34]

def extrair_numero_nf(texto, chaves=None):
    chaves = chaves or []
    for chave in chaves:
        if validar_chave_nfe(chave):
            return numero_nf_da_chave(chave)
    for m in REGEX_NUMERO_NF.finditer(str(texto or "")):
        numero = somente_digitos(m.group(1)).lstrip("0")
        if numero:
            return numero
    return ""

def extrair_urls_consulta_nf(texto):
    urls = []
    for url in REGEX_URL.findall(str(texto or "")):
        url_limpa = url.rstrip('.,;)]}')
        url_norm = normalizar(url_limpa)
        if any(termo in url_norm for termo in ["nfe", "nfce", "nfse", "sefaz", "sped", "fazenda", "consulta", "qrcode"]):
            if url_limpa not in urls:
                urls.append(url_limpa)
    return urls

def extrair_valor_nota(texto):
    texto = limpar_texto(texto)
    rotulos = [
        "valor total da nota", "valor total", "valor pago", "valor emitido",
        "valor do documento", "total da nota", "total dos serviços", "total dos servicos",
    ]
    texto_norm = normalizar(texto)
    for rotulo in rotulos:
        pos = texto_norm.find(normalizar(rotulo))
        if pos != -1:
            trecho = texto[pos:pos + 180]
            valor = escolher_valor_monetario(trecho)
            if valor is not None:
                return abs(valor)
    valor = escolher_valor_monetario(texto)
    return abs(valor) if valor is not None else None

def consultar_link_nf(url):
    if not url or os.getenv("CONSULTAR_NF_ONLINE", "1") == "0":
        return "Consulta online não executada"
    try:
        req = Request(url, headers={"User-Agent": "Mozilla/5.0 AuditoriaCondominio/1.0"})
        with urlopen(req, timeout=10) as resp:
            return f"Link consultado com sucesso (HTTP {resp.status})"
    except HTTPError as exc:
        return f"Link consultado, retorno HTTP {exc.code}"
    except (URLError, TimeoutError, ValueError) as exc:
        return f"Não foi possível consultar o link automaticamente: {exc}"

def aguardar_download_nf():
    global _NF_LAST_DOWNLOAD
    decorrido = time.time() - _NF_LAST_DOWNLOAD
    if decorrido < NF_DOWNLOAD_DELAY:
        time.sleep(NF_DOWNLOAD_DELAY - decorrido)
    _NF_LAST_DOWNLOAD = time.time()

def decodificar_bytes_http(data, content_type=""):
    charset = ""
    m = re.search(r"charset=([^;]+)", str(content_type), re.IGNORECASE)
    if m:
        charset = m.group(1).strip()
    for enc in [charset, "utf-8", "latin1"]:
        if not enc:
            continue
        try:
            return data.decode(enc, errors="ignore")
        except Exception:
            continue
    return data.decode("latin1", errors="ignore")

def extrair_links_download_html(html_texto, base_url):
    links = []
    candidatos = re.findall(r'(?:href|src)=["\']([^"\']+)', html_texto, re.IGNORECASE)
    candidatos.extend(REGEX_URL.findall(html_texto))
    for candidato in candidatos:
        url = urljoin(base_url, candidato.strip())
        url_norm = normalizar(unquote_plus(url))
        if any(k in url_norm for k in ["download", "baixar", "arquivo", "anexo", "documento", ".pdf"]):
            if url not in links:
                links.append(url)
    return links

def salvar_bytes_documento(url, data, content_type, prefixo):
    parsed = urlparse(url)
    nome_url = Path(parsed.path).name
    ext = Path(nome_url).suffix.lower()
    if not ext:
        if data[:4] == b"%PDF" or "pdf" in str(content_type).lower():
            ext = ".pdf"
        elif "html" in str(content_type).lower():
            ext = ".html"
        elif "png" in str(content_type).lower():
            ext = ".png"
        elif "jpeg" in str(content_type).lower() or "jpg" in str(content_type).lower():
            ext = ".jpg"
        else:
            ext = ".bin"
    nome = f"{prefixo}_{gerar_hash(url)[:12]}{ext}"
    caminho = Path(PASTA_DOCUMENTOS_NF) / nome
    if not caminho.exists():
        with open(caminho, "wb") as f:
            f.write(data)
    return str(caminho)

def baixar_documento_nf(url, prefixo="nf"):
    global _NF_DOWNLOAD_COUNT
    if not url:
        return {"arquivo": "", "status": "Sem link para baixar", "url_final": ""}
    if url in _NF_DOWNLOAD_CACHE:
        return _NF_DOWNLOAD_CACHE[url]
    if NF_MAX_DOWNLOADS > 0 and _NF_DOWNLOAD_COUNT >= NF_MAX_DOWNLOADS:
        return {"arquivo": "", "status": f"Limite de {NF_MAX_DOWNLOADS} downloads atingido nesta execução", "url_final": url}

    try:
        aguardar_download_nf()
        _NF_DOWNLOAD_COUNT += 1
        req = Request(url, headers={"User-Agent": "Mozilla/5.0 AuditoriaCondominio/1.0"})
        with urlopen(req, timeout=NF_DOWNLOAD_TIMEOUT) as resp:
            data = resp.read(NF_MAX_BYTES + 1)
            content_type = resp.headers.get("Content-Type", "")
            url_final = resp.geturl()
        if len(data) > NF_MAX_BYTES:
            resultado = {"arquivo": "", "status": "Documento excede limite de tamanho configurado", "url_final": url_final}
        elif data[:4] == b"%PDF" or "pdf" in content_type.lower():
            arquivo = salvar_bytes_documento(url_final, data, content_type, prefixo)
            resultado = {"arquivo": arquivo, "status": "PDF baixado do link do balancete", "url_final": url_final}
        elif "html" in content_type.lower() or data.lstrip().lower().startswith((b"<!doctype", b"<html")):
            html_texto = decodificar_bytes_http(data, content_type)
            arquivo_html = salvar_bytes_documento(url_final, data, content_type, prefixo)
            for candidato in extrair_links_download_html(html_texto, url_final)[:6]:
                if candidato == url:
                    continue
                baixado = baixar_documento_nf(candidato, prefixo)
                if baixado.get("arquivo") and baixado["arquivo"].lower().endswith(".pdf"):
                    resultado = {**baixado, "status": baixado.get("status", "") + " via página intermediária", "url_final": baixado.get("url_final", candidato)}
                    _NF_DOWNLOAD_CACHE[url] = resultado
                    return resultado
            resultado = {"arquivo": arquivo_html, "status": "Página do link salva; PDF anexo não localizado automaticamente", "url_final": url_final}
        else:
            arquivo = salvar_bytes_documento(url_final, data, content_type, prefixo)
            resultado = {"arquivo": arquivo, "status": f"Arquivo baixado ({content_type or 'tipo desconhecido'})", "url_final": url_final}
    except HTTPError as exc:
        resultado = {"arquivo": "", "status": f"Falha ao baixar link: HTTP {exc.code}", "url_final": url}
    except (URLError, TimeoutError, ValueError) as exc:
        resultado = {"arquivo": "", "status": f"Falha ao baixar link lentamente: {exc}", "url_final": url}
    _NF_DOWNLOAD_CACHE[url] = resultado
    return resultado

def texto_pdf_local(caminho_pdf, max_paginas=4):
    try:
        doc = fitz.open(caminho_pdf)
        textos = []
        for idx in range(min(len(doc), max_paginas)):
            pagina = doc[idx]
            texto = pagina.get_text("text")
            if len(texto.strip()) < 30:
                texto = executar_ocr(pagina)
            textos.append(texto)
        return "\n".join(textos)
    except Exception:
        return ""

def imagem_primeira_pagina_pdf(caminho_pdf, prefixo):
    try:
        doc = fitz.open(caminho_pdf)
        if len(doc) == 0:
            return ""
        pix = doc[0].get_pixmap(matrix=fitz.Matrix(2, 2))
        caminho = os.path.join(PASTA_IMAGENS, f"{prefixo}_{gerar_hash(caminho_pdf)[:12]}.png")
        if not os.path.exists(caminho):
            pix.save(caminho, output="png")
        return caminho
    except Exception:
        return ""

def extrair_nota_de_link(url, valor, descricao):
    prefixo = nome_arquivo_seguro(f"link_{valor}_{descricao}", 60)
    baixado = baixar_documento_nf(url, prefixo)
    arquivo = baixado.get("arquivo", "")
    texto = ""
    imagem = ""
    if arquivo.lower().endswith(".pdf"):
        texto = texto_pdf_local(arquivo)
        imagem = imagem_primeira_pagina_pdf(arquivo, prefixo)
    elif arquivo.lower().endswith((".png", ".jpg", ".jpeg")):
        imagem = arquivo
        try:
            texto = pytesseract.image_to_string(Image.open(arquivo), lang="por")
        except Exception:
            texto = ""
    elif arquivo.lower().endswith(".html"):
        try:
            with open(arquivo, "rb") as f:
                texto = decodificar_bytes_http(f.read(), "text/html")
        except Exception:
            texto = ""

    info = montar_info_nota(texto, imagem, valor)
    info["status_consulta_nf"] = baixado.get("status", "") if not info.get("numero_nf") and not info.get("chave_nf") else f"{baixado.get('status', '')}; metadados extraídos do documento"
    info["link_consulta_nf"] = info.get("link_consulta_nf") or baixado.get("url_final", "")
    info["link_origem_nf"] = url
    info["arquivo_nf_baixado"] = arquivo
    return info

def montar_info_nota(texto, caminho_imagem, valor):
    chaves = extrair_chaves_nfe(texto)
    chave_valida = next((chave for chave in chaves if validar_chave_nfe(chave)), "")
    chave_invalida = chaves[0] if chaves and not chave_valida else ""
    numero_nf = extrair_numero_nf(texto, chaves)
    urls = extrair_urls_consulta_nf(texto)
    link_consulta = urls[0] if urls else ""
    if link_consulta:
        status = consultar_link_nf(link_consulta)
    elif chave_valida:
        status = "Chave NF-e válida; consulta oficial exige portal/SEFAZ com CAPTCHA ou certificado"
    elif chave_invalida:
        status = "Chave/código fiscal encontrado; validação NF-e automática inconclusiva"
    elif numero_nf:
        status = "Número de NF encontrado; chave de acesso não localizada para consulta automática"
    else:
        status = "Imagem vinculada por valor; número/chave de NF não localizado no OCR"
    return {
        "valor": round(abs(valor), 2) if valor is not None else None,
        "imagem": caminho_imagem,
        "numero_nf": numero_nf,
        "chave_nf": chave_valida or chave_invalida,
        "status_consulta_nf": status,
        "link_consulta_nf": link_consulta,
        "link_origem_nf": "",
        "arquivo_nf_baixado": "",
    }

def extrair_notas(pdf_path):
    doc = fitz.open(pdf_path)
    mapa = {}
    for pagina_num in range(len(doc)):
        pagina = doc[pagina_num]
        texto = pagina.get_text("text")
        if len(texto.strip()) < 30: texto = executar_ocr(pagina)
        texto_lower = normalizar(texto)
        termos_nf = ["nota fiscal", "nf-e", "nfe", "nfs-e", "nfse", "danfe", "chave de acesso", "documento auxiliar", "cupom fiscal", "qrcode"]
        termos_pagamento = ["pix", "comprovante", "darf", "boleto", "pago a", "receita federal", "transferencia", "pagamento", "valor pago", "valor emitido", "valor total"]
        if not any(termo in texto_lower for termo in termos_nf + termos_pagamento): continue
        valor = extrair_valor_nota(texto)
        if valor is None: continue
        try:
            pix = pagina.get_pixmap(matrix=fitz.Matrix(2, 2))
            valor_limpo = str(abs(valor)).replace(".", "_").replace("-", "")
            numero_nf = extrair_numero_nf(texto, extrair_chaves_nfe(texto))
            sufixo_nf = f"_nf_{numero_nf}" if numero_nf else ""
            nome = f"NF_pag_{pagina_num + 1:03d}_valor_{valor_limpo}{sufixo_nf}.png"
            caminho = os.path.join(PASTA_IMAGENS, nome)
            pix.save(caminho, output="png")
            info = montar_info_nota(texto, caminho, valor)
            mapa.setdefault(round(abs(valor), 2), []).append(info)
        except Exception: continue
    return mapa

def limpar_linha_movimentacao(linha):
    linha = str(linha).replace("**", " ")
    linha = re.sub(r"Rua Professor.*|Emitido em.*|\[.*?\]|https?://\S+", "", linha, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", linha).strip()

def identificar_conta_pagina(texto):
    m = re.search(r'Movimentação Analítica\s+"([^"]+)"|Movimentacao Analitica\s+"([^"]+)"', texto, re.IGNORECASE)
    return m.group(1).strip() if m and m.group(1) else (m.group(2).strip() if m and m.group(2) else "")

def tipo_movimento_por_valor(valor, descricao):
    desc_norm = normalizar(descricao)
    if "transf." in desc_norm or "transferencia" in desc_norm or "para a conta" in desc_norm: return "Transferência"
    if valor > 0: return "Entrada"
    if valor < 0: return "Saída"
    return "Neutro"

def obter_info_nota_por_valor(mapa_notas, valor):
    chave = round(abs(valor), 2)
    notas = mapa_notas.get(chave, [])
    if isinstance(notas, dict):
        return notas
    if notas:
        return notas.pop(0) if len(notas) > 1 else notas[0]
    return {}

def tokens_relevantes(texto):
    ignorar = {"para", "com", "das", "dos", "mes", "condominio", "condominial", "pagamento", "compra", "servico", "servicos", "realizado", "realizada", "referente"}
    return [t for t in re.findall(r"[a-z0-9]{4,}", normalizar(unquote_plus(str(texto or "")))) if t not in ignorar]

def escolher_link_movimentacao(descricao, fornecedor, links):
    if not links:
        return ""
    desc_tokens = set(tokens_relevantes(descricao))
    fornecedor_tokens = set(tokens_relevantes(fornecedor))
    melhor_link, melhor_score = "", 0
    for link in links:
        link_decod = normalizar(unquote_plus(link))
        link_tokens = set(tokens_relevantes(link_decod))
        score = len(desc_tokens & link_tokens) * 4 + len(fornecedor_tokens & link_tokens) * 10
        if fornecedor and normalizar(fornecedor) in link_decod:
            score += 30
        if "itemrelacionado" in link_decod:
            score += 8
        if any(k in link_decod for k in ["despesa", "arquivo", "documento"]):
            score += 5
        if score > melhor_score:
            melhor_link, melhor_score = link, score
    return melhor_link if melhor_score >= 12 else ""

def score_suspeita_movimentacao(valor, descricao, nota_info):
    valor_abs, desc_norm, score, motivos = abs(valor), normalizar(descricao), 0, []
    link_nf = nota_info.get("imagem", "") if isinstance(nota_info, dict) else ""
    status_nf = normalizar(nota_info.get("status_consulta_nf", "")) if isinstance(nota_info, dict) else ""
    if valor >= 0: return 0, ""
    if valor_abs > 5000:
        score += 20
        motivos.append(f"Valor alto ({float_para_br(valor_abs)})")
    if valor_abs > 500 and not link_nf:
        score += 25
        motivos.append(f"Sem documento vinculado para saída de {float_para_br(valor_abs)}")
    if detectar_pix(descricao):
        score += 12
        motivos.append("Pagamento via PIX/TED/CPF exige conferência do favorecido")
    if "sem apresentacao" in desc_norm:
        score += 50
        motivos.append("Descrição informa ausência de apresentação de NF")
    if link_nf and ("invalido" in status_nf or "não localizado" in status_nf or "nao localizado" in status_nf):
        score += 10
        motivos.append("Documento vinculado sem chave/número validado")
    return min(score, 100), " | ".join(motivos)

def recalcular_scores_divergencia(df_mov):
    if df_mov.empty or "Tipo_Movimento" not in df_mov.columns:
        return df_mov
    out = df_mov.copy()
    out.loc[out["Tipo_Movimento"] != "Saída", "Score_Suspeita"] = 0
    out.loc[out["Tipo_Movimento"] != "Saída", "Motivo_Suspeita"] = ""
    saidas = out[out["Tipo_Movimento"] == "Saída"].copy()
    if saidas.empty:
        return out
    saidas["Valor_Real"] = pd.to_numeric(saidas["Valor_Real"], errors="coerce").fillna(0)
    p90_mes = saidas.groupby("Mes_Ano")["Valor_Real"].quantile(0.90).to_dict() if "Mes_Ano" in saidas.columns else {}
    p95_geral = float(saidas["Valor_Real"].quantile(0.95)) if len(saidas) else 0
    duplicados = saidas.groupby(["Mes_Ano", "Fornecedor", "Valor_Real"], dropna=False).size().to_dict() if {"Mes_Ano", "Fornecedor", "Valor_Real"}.issubset(saidas.columns) else {}

    for idx, row in saidas.iterrows():
        valor = float(row.get("Valor_Real", 0) or 0)
        desc = str(row.get("Descricao", ""))
        desc_norm = normalizar(desc)
        fornecedor = str(row.get("Fornecedor", "") or "").strip()
        categoria = str(row.get("Categoria", "") or "").strip()
        tem_doc = bool(str(row.get("Link_NF", "") or "").strip())
        tem_download = bool(str(row.get("Arquivo_NF_Baixado", "") or "").strip())
        numero_nf = str(row.get("Numero_NF", "") or "").strip()
        chave_nf = str(row.get("Chave_NF", "") or "").strip()
        status_nf = normalizar(row.get("Status_Consulta_NF", ""))
        score, motivos = 0, []

        base_mes = max(float(p90_mes.get(row.get("Mes_Ano"), 0) or 0), p95_geral * 0.65)
        if valor >= 5000 and valor >= base_mes:
            score += 28
            motivos.append(f"Valor {float_para_br(valor)} acima do padrão do período")
        elif valor >= 2000:
            score += 12
            motivos.append(f"Valor relevante ({float_para_br(valor)})")

        if not tem_doc and valor >= 500:
            score += 30 if valor >= 2000 else 18
            motivos.append("Não há nota/comprovante vinculado ao lançamento")
        elif tem_doc and not (numero_nf or chave_nf):
            score += 8
            motivos.append("Documento vinculado, mas OCR não identificou número/chave da nota")
        if tem_download:
            score = max(0, score - 8)

        if len(somente_digitos(chave_nf)) == 44 and validar_chave_nfe(chave_nf):
            score = max(0, score - 5)
        if any(k in status_nf for k in ["falha", "limite", "não foi possível", "nao foi possivel"]):
            score += 8
            motivos.append("Consulta/download do documento não foi confirmada automaticamente")

        if detectar_pix(desc):
            score += 12
            motivos.append("PIX/TED/CPF na descrição; conferir favorecido e comprovante")
        if any(k in desc_norm for k in ["sem apresentacao", "sem nf", "sem nota"]):
            score += 45
            motivos.append("Descrição indica ausência de nota fiscal")
        if categoria in ["Outros", "Manutenção/Obras"] and not str(row.get("CNPJ", "") or "").strip() and valor >= 500:
            score += 10
            motivos.append(f"Categoria {categoria} sem CNPJ identificado na descrição")
        if not fornecedor and valor >= 500:
            score += 8
            motivos.append("Fornecedor não identificado automaticamente")

        chave_dup = (row.get("Mes_Ano"), row.get("Fornecedor"), valor)
        if duplicados.get(chave_dup, 0) >= 2 and valor >= 300:
            score += 8
            motivos.append("Mesmo fornecedor/valor aparece mais de uma vez no mês")

        out.at[idx, "Score_Suspeita"] = min(int(score), 100)
        out.at[idx, "Motivo_Suspeita"] = " | ".join(motivos)
    return out

def parse_movimentacoes_texto(texto):
    registros = []
    texto_limpo = re.sub(r"https?://\S+|\[.*?\]|Rua Professor.*|Contatos:.*|CRA MG:.*", " ", str(texto).replace("\r", " "), flags=re.IGNORECASE)
    linhas = [l.strip() for l in texto_limpo.split("\n") if l.strip()]
    
    # Encontrar posição do cabeçalho "Saldo" e cortar após ele
    idx_inicio = 0
    for idx, linha in enumerate(linhas):
        if normalizar(linha) == "saldo":
            idx_inicio = idx + 1
            break
    
    # Pular saldo inicial (primeiro valor após cabeçalho)
    while idx_inicio < len(linhas):
        if re.match(REGEX_VALOR_TXT, linhas[idx_inicio].replace("(", "").replace(")", "").replace("-", "")):
            idx_inicio += 1
            break
        elif re.match(r"^\d{2}/\d{2}/\d{4}$", linhas[idx_inicio]):
            break
        else:
            idx_inicio += 1
    
    linhas = linhas[idx_inicio:]

    i = 0
    while i < len(linhas):
        linha = linhas[i]
        data_match = re.match(r"^(\d{2}/\d{2}/\d{4})$", linha.strip())
        if data_match:
            data = data_match.group(1)
            descricao_partes = []
            valor_txt = ""
            saldo_txt = ""
            j = i + 1
            while j < len(linhas) and j < i + 8:
                cand = linhas[j].strip()
                prox_data = re.match(r"^\d{2}/\d{2}/\d{4}$", cand)
                if prox_data:
                    break
                if re.match(REGEX_VALOR_TXT, cand.replace("(", "").replace(")", "").replace("-", "")):
                    if not valor_txt:
                        valor_txt = cand
                    else:
                        saldo_txt = cand
                        break
                elif cand and len(cand) > 1 and not cand.startswith("-"):
                    descricao_partes.append(cand)
                j += 1
            descricao = " ".join(descricao_partes) if descricao_partes else ""
            if descricao and valor_txt:
                valor = valor_para_float(valor_txt)
                saldo = valor_para_float(saldo_txt) if saldo_txt else None
                if valor is not None:
                    registros.append({"Data": data, "Descricao": limpar_linha_movimentacao(descricao), "Valor_Movimento": valor, "Saldo_Apos": saldo})
            i = j if j > i else i + 1
        else:
            i += 1
    return registros

def agrupar_paginas_movimentacao(paginas):
    """Agrupa páginas consecutivas que pertencem à mesma movimentação analítica"""
    grupos = []
    grupo_atual = []
    conta_atual = ""
    
    for pagina in paginas:
        texto = pagina["texto"]
        tem_mov = "movimentacao analitica" in normalizar(texto)
        
        if tem_mov:
            # Nova conta encontrada
            conta = identificar_conta_pagina(texto)
            if grupo_atual and conta_atual:
                grupos.append({"conta": conta_atual, "paginas": grupo_atual})
            grupo_atual = [pagina]
            conta_atual = conta
        elif grupo_atual:
            # Verifica se é continuação da movimentação anterior
            linhas = [l.strip() for l in texto.split('\n') if l.strip()]
            tem_formato = any(re.match(r'^\d{2}/\d{2}/\d{4}$', l) for l in linhas)
            if tem_formato:
                grupo_atual.append(pagina)
            else:
                # Fim do grupo atual
                if grupo_atual and conta_atual:
                    grupos.append({"conta": conta_atual, "paginas": grupo_atual})
                grupo_atual = []
                conta_atual = ""
    
    # Adicionar último grupo
    if grupo_atual and conta_atual:
        grupos.append({"conta": conta_atual, "paginas": grupo_atual})
    
    return grupos

def extrair_movimentacoes_analiticas(pdf_path, paginas, meta, mapa_notas):
    registros = []
    grupos = agrupar_paginas_movimentacao(paginas)
    
    # Contas que não são operacionais (não entram no balancete de receitas/despesas)
    contas_excluidas = ["conta capital"]
    
    for grupo in grupos:
        conta = grupo["conta"]
        # Pular contas não operacionais
        if conta.lower() in contas_excluidas:
            continue
            
        for pagina in grupo["paginas"]:
            for mov in parse_movimentacoes_texto(pagina["texto"]):
                data_dt = converter_data(mov["Data"])
                ano, mes, mes_ano = (int(data_dt.year), int(data_dt.month), data_dt.strftime("%Y-%m")) if pd.notna(data_dt) else (meta.get("Ano"), meta.get("Mes"), meta.get("Mes_Ano"))
                tipo = tipo_movimento_por_valor(mov["Valor_Movimento"], mov["Descricao"])
                fornecedor = extrair_fornecedor(mov["Descricao"])
                nota_info = obter_info_nota_por_valor(mapa_notas, mov["Valor_Movimento"])
                link_origem = escolher_link_movimentacao(mov["Descricao"], fornecedor, pagina.get("links", [])) if tipo == "Saída" else ""
                if link_origem and not nota_info.get("arquivo_nf_baixado"):
                    nota_link = extrair_nota_de_link(link_origem, abs(mov["Valor_Movimento"]), mov["Descricao"])
                    if nota_link.get("imagem") or nota_link.get("arquivo_nf_baixado") or not nota_info:
                        nota_info = {**nota_info, **{k: v for k, v in nota_link.items() if v}}
                score, motivos = score_suspeita_movimentacao(mov["Valor_Movimento"], mov["Descricao"], nota_info)
                registros.append({
                    **meta, "Data": mov["Data"], "Ano": ano, "Mes": mes, "Mes_Ano": mes_ano, "Conta": conta, "Tipo_Movimento": tipo,
                    "Categoria": classificar_categoria(mov["Descricao"]), "Fornecedor": fornecedor,
                    "Descricao": mov["Descricao"], "Valor_Real": abs(mov["Valor_Movimento"]), "Valor_Assinado": mov["Valor_Movimento"],
                    "Valor_Entrada": abs(mov["Valor_Movimento"]) if tipo == "Entrada" else 0, "Valor_Saida": abs(mov["Valor_Movimento"]) if tipo == "Saída" else 0,
                    "Valor_Transferencia": abs(mov["Valor_Movimento"]) if tipo == "Transferência" else 0, "Saldo_Apos": mov["Saldo_Apos"],
                    "CNPJ": extrair_cnpj(mov["Descricao"]), "PIX": detectar_pix(mov["Descricao"]), "Score_Suspeita": score,
                    "Motivo_Suspeita": motivos, "Numero_NF": nota_info.get("numero_nf", ""), "Chave_NF": nota_info.get("chave_nf", ""),
                    "Status_Consulta_NF": nota_info.get("status_consulta_nf", ""), "Link_Consulta_NF": nota_info.get("link_consulta_nf", ""),
                    "Link_Origem_NF": nota_info.get("link_origem_nf", link_origem), "Arquivo_NF_Baixado": nota_info.get("arquivo_nf_baixado", ""),
                    "Link_NF": nota_info.get("imagem", ""), "Origem_Extracao": "Movimentação Analítica",
                    "Hash": gerar_hash(f"{pdf_path.name}_{conta}_{mov['Data']}_{mov['Descricao']}_{mov['Valor_Movimento']}_{mov['Saldo_Apos']}")
                })
    return registros

def extrair_composicao_cobrancas(pdf_path, paginas, meta):
    registros, unidade_atual, bloco_atual = [], "", ""
    for pagina in paginas:
        if "Composição das cobranças por crédito" not in pagina["texto"]: continue
        for linha in pagina["linhas"]:
            linha = limpar_linha_movimentacao(linha)
            m_unidade = re.match(r"^(\d{3})\s+(\d{2})\s+Cobrança", linha, re.IGNORECASE)
            if m_unidade:
                unidade_atual, bloco_atual = m_unidade.group(1), m_unidade.group(2)
                continue
            m = re.match(r"^(\d{5,6})\s+(.+?)\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+(" + REGEX_VALOR_TXT + r")\s*$", linha)
            if not m: continue
            valor = valor_para_float(m.group(6))
            if valor is None: continue
            data_dt = converter_data(m.group(5))
            ano, mes, mes_ano = (int(data_dt.year), int(data_dt.month), data_dt.strftime("%Y-%m")) if pd.notna(data_dt) else (meta.get("Ano"), meta.get("Mes"), meta.get("Mes_Ano"))
            registros.append({
                **meta, "Ano": ano, "Mes": mes, "Mes_Ano": mes_ano, "Unidade": unidade_atual, "Bloco": bloco_atual,
                "Codigo_Cobranca": m.group(1), "Descricao": limpar_linha(m.group(2)), "Categoria": classificar_categoria(limpar_linha(m.group(2))),
                "Vencimento": m.group(3), "Data_Credito": m.group(4), "Data_Liquidacao": m.group(5), "Valor": abs(valor),
                "Valor_Assinado": valor, "Tipo_Movimento": "Entrada" if valor >= 0 else "Saída", "Origem_Extracao": "Composição das Cobranças",
                "Hash": gerar_hash(f"{pdf_path.name}_{unidade_atual}_{bloco_atual}_{m.group(1)}_{limpar_linha(m.group(2))}_{m.group(3)}_{m.group(4)}_{m.group(5)}_{valor}")
            })
    return registros

# ============================================================
# VALIDAÇÃO, FRAUDES, DASHBOARDS E EXPORTAÇÃO
# ============================================================

def criar_validacao(df_balanco, df_mov):
    registros = []
    if df_balanco.empty: return pd.DataFrame(columns=COLUNAS_VALIDACAO)
    if df_mov.empty or "Arquivo_Origem" not in df_mov.columns:
        for _, row in df_balanco.iterrows():
            registros.append({"Arquivo_Origem": row.get("Arquivo_Origem"), "Mes_Ano": row.get("Mes_Ano"), "Receitas_Balancete": row.get("Total_Receitas"), "Entradas_Movimentacao": 0, "Diferenca_Entradas": None, "Despesas_Balancete": row.get("Total_Despesas"), "Saidas_Movimentacao": 0, "Diferenca_Saidas": None, "Movimento_Liquido_Balancete": row.get("Movimento_Liquido"), "Saldo_Anterior": row.get("Saldo_Anterior"), "Saldo_Final": row.get("Saldo_Final"), "Status": "SEM_MOVIMENTAÇÃO_ANALÍTICA_EXTRAÍDA"})
        return pd.DataFrame(registros, columns=COLUNAS_VALIDACAO)

    for _, row in df_balanco.iterrows():
        mov_mes = df_mov[(df_mov["Arquivo_Origem"] == row.get("Arquivo_Origem")) & (df_mov["Mes_Ano"] == row.get("Mes_Ano"))]
        entradas_mov = mov_mes["Valor_Entrada"].sum() if not mov_mes.empty else 0
        saidas_mov = mov_mes["Valor_Saida"].sum() if not mov_mes.empty else 0
        diff_entradas = round(entradas_mov - row.get("Total_Receitas", 0), 2) if pd.notna(row.get("Total_Receitas")) else None
        diff_saidas = round(saidas_mov - row.get("Total_Despesas", 0), 2) if pd.notna(row.get("Total_Despesas")) else None
        status = "VERIFICAR_DIFERENÇA_MOVIMENTAÇÃO_X_BALANCETE" if diff_entradas not in [None, 0] or diff_saidas not in [None, 0] else "OK"
        registros.append({"Arquivo_Origem": row.get("Arquivo_Origem"), "Mes_Ano": row.get("Mes_Ano"), "Receitas_Balancete": row.get("Total_Receitas"), "Entradas_Movimentacao": entradas_mov, "Diferenca_Entradas": diff_entradas, "Despesas_Balancete": row.get("Total_Despesas"), "Saidas_Movimentacao": saidas_mov, "Diferenca_Saidas": diff_saidas, "Movimento_Liquido_Balancete": row.get("Movimento_Liquido"), "Saldo_Anterior": row.get("Saldo_Anterior"), "Saldo_Final": row.get("Saldo_Final"), "Status": status})
    return pd.DataFrame(registros, columns=COLUNAS_VALIDACAO)

def detectar_fraudes(df_mov):
    if df_mov.empty or "Tipo_Movimento" not in df_mov.columns or "Score_Suspeita" not in df_mov.columns: return pd.DataFrame(columns=COLUNAS_MOVIMENTACOES)
    return df_mov[(df_mov["Tipo_Movimento"] == "Saída") & (df_mov["Score_Suspeita"] >= 40)].copy()

def criar_dashboards(df_balanco, df_cat, df_mov):
    if not df_balanco.empty:
        bal = df_balanco.copy().sort_values("Mes_Ano")
        bal_melt = bal.melt(id_vars=["Mes_Ano"], value_vars=["Total_Receitas", "Total_Despesas", "Movimento_Liquido"], var_name="Indicador", value_name="Valor")
        bal_melt["Indicador"] = bal_melt["Indicador"].replace({"Total_Receitas": "Entradas do mês", "Total_Despesas": "Saídas do mês", "Movimento_Liquido": "Balanço do mês"})
        px.bar(bal_melt, x="Mes_Ano", y="Valor", color="Indicador", barmode="group", title="Entradas, Saídas e Balanço do Mês").write_html("dashboard_balanco_mensal.html")
        
        fig_saldo = go.Figure()
        fig_saldo.add_trace(go.Scatter(x=bal["Mes_Ano"], y=bal["Saldo_Anterior"], mode="lines+markers", name="Saldo anterior"))
        fig_saldo.add_trace(go.Scatter(x=bal["Mes_Ano"], y=bal["Saldo_Final"], mode="lines+markers", name="Saldo final"))
        fig_saldo.update_layout(title="Saldo Anterior x Saldo Final", xaxis_title="Mês", yaxis_title="Valor").write_html("dashboard_saldos.html")

    if not df_cat.empty:
        cat = df_cat.copy()
        px.bar(cat, x="Categoria", y="Valor", color="Grupo", barmode="group", title="Categorias de Entradas e Saídas").write_html("dashboard_categorias_entrada_saida.html")
        px.bar(cat, x="Mes_Ano", y="Valor", color="Categoria", facet_col="Grupo", title="Categorias por Mês - Entradas e Saídas").write_html("dashboard_categorias_por_mes.html")

    if not df_mov.empty and "Tipo_Movimento" in df_mov.columns:
        mov = df_mov[df_mov["Tipo_Movimento"].isin(["Entrada", "Saída", "Transferência"])].copy()
        if not mov.empty:
            mov_group = mov.groupby(["Mes_Ano", "Conta", "Tipo_Movimento"])[["Valor_Entrada", "Valor_Saida", "Valor_Transferencia"]].sum().reset_index()
            mov_group["Valor"] = mov_group.apply(lambda row: row["Valor_Entrada"] if row["Tipo_Movimento"] == "Entrada" else (row["Valor_Saida"] if row["Tipo_Movimento"] == "Saída" else row["Valor_Transferencia"]), axis=1)
            px.bar(mov_group, x="Mes_Ano", y="Valor", color="Tipo_Movimento", facet_col="Conta", title="Movimentação Analítica por Conta").write_html("dashboard_movimentacao_por_conta.html")
            
            top_saida = mov[mov["Tipo_Movimento"] == "Saída"].groupby("Fornecedor")["Valor_Saida"].sum().sort_values(ascending=False).head(15).reset_index()
            if not top_saida.empty: px.bar(top_saida, x="Fornecedor", y="Valor_Saida", title="Top Fornecedores por Saída").write_html("dashboard_top_fornecedores_saida.html")

        if "Link_NF" in df_mov.columns:
            notas = df_mov[df_mov["Link_NF"].fillna("").astype(str).str.len() > 0].copy()
            if not notas.empty and "Status_Consulta_NF" in notas.columns:
                status_nf = notas["Status_Consulta_NF"].fillna("Sem status").astype(str).value_counts().reset_index()
                status_nf.columns = ["Status", "Quantidade"]
                px.bar(status_nf, x="Quantidade", y="Status", orientation="h", title="Status das Notas Fiscais e Comprovantes").write_html("dashboard_notas_fiscais.html")

def exportar_excel(df_balanco, df_cat, df_mov, df_cobrancas, df_fraudes, df_validacao):
    with pd.ExcelWriter(ARQUIVO_EXCEL, engine="openpyxl") as writer:
        df_balanco.to_excel(writer, sheet_name="Balanco_Mensal", index=False)
        df_cat.to_excel(writer, sheet_name="Categorias_Mensais", index=False)
        df_mov.to_excel(writer, sheet_name="Movimentacoes", index=False)
        if not df_mov.empty and "Tipo_Movimento" in df_mov.columns:
            df_mov[df_mov["Tipo_Movimento"] == "Entrada"].to_excel(writer, sheet_name="Entradas_Analiticas", index=False)
            df_mov[df_mov["Tipo_Movimento"] == "Saída"].to_excel(writer, sheet_name="Saidas_Analiticas", index=False)
            df_mov[df_mov["Tipo_Movimento"] == "Transferência"].to_excel(writer, sheet_name="Transferencias", index=False)
        df_cobrancas.to_excel(writer, sheet_name="Composicao_Cobrancas", index=False)
        df_fraudes.to_excel(writer, sheet_name="Fraudes", index=False)
        df_validacao.to_excel(writer, sheet_name="Validacao", index=False)
        if not df_mov.empty and "Link_NF" in df_mov.columns:
            df_notas = df_mov[df_mov["Link_NF"].fillna("").astype(str).str.len() > 0].copy()
            if not df_notas.empty:
                colunas_notas = [c for c in ["Arquivo_Origem", "Mes_Ano", "Data", "Fornecedor", "Descricao", "Valor_Real", "Numero_NF", "Chave_NF", "Status_Consulta_NF", "Link_Consulta_NF", "Link_Origem_NF", "Arquivo_NF_Baixado", "Link_NF"] if c in df_notas.columns]
                df_notas[colunas_notas].to_excel(writer, sheet_name="Notas_Fiscais", index=False)
    
    wb = load_workbook(ARQUIVO_EXCEL)
    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        for cell in ws[1]:
            cell.font, cell.fill = Font(bold=True, color="FFFFFF"), PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.freeze_panes = "A2"
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(max((len(str(c.value)) for c in col if c.value is not None), default=0) + 2, 70)
        
        if nome_aba == "Movimentacoes":
            colunas = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
            if "Link_NF" in colunas:
                col_nf, coluna_img = colunas["Link_NF"], ws.max_column + 1
                ws.cell(row=1, column=coluna_img, value="Miniatura_NF").font = Font(bold=True, color="FFFFFF")
                ws.cell(row=1, column=coluna_img).fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                for linha in range(2, ws.max_row + 1):
                    caminho = ws.cell(linha, col_nf).value
                    if caminho and os.path.exists(caminho):
                        try:
                            img = XLImage(caminho)
                            img.width, img.height = 70, 70
                            ws.add_image(img, ws.cell(row=linha, column=coluna_img).coordinate)
                            ws.row_dimensions[linha].height = 60
                        except Exception: pass
    wb.save(ARQUIVO_EXCEL)

def exportar_csvs(df_balanco, df_cat, df_mov, df_cobrancas):
    for df, arq in [(df_balanco, ARQUIVO_CSV_BALANCO), (df_cat, ARQUIVO_CSV_CATEGORIAS), (df_mov, ARQUIVO_CSV_MOVIMENTACOES), (df_cobrancas, ARQUIVO_CSV_COBRANCAS)]:
        df.to_csv(arq, sep=";", index=False, encoding="utf-8-sig")

def processar_pdf(pdf_path):
    paginas = extrair_paginas_pdf(pdf_path)
    meta = extrair_metadados(pdf_path, paginas)
    mapa_notas = extrair_notas(pdf_path)
    return extrair_balanco_mensal(pdf_path, paginas, meta), extrair_categorias_mensais(pdf_path, paginas, meta), extrair_movimentacoes_analiticas(pdf_path, paginas, meta, mapa_notas), extrair_composicao_cobrancas(pdf_path, paginas, meta)

def main(progress_callback=None):
    pdfs = list(Path(PASTA_PDFS).glob("*.pdf"))
    if not pdfs: return
    if progress_callback:
        progress_callback(0, f'Limpando arquivos antigos... ({len(pdfs)} PDFs encontrados)')
    for imagem_antiga in Path(PASTA_IMAGENS).glob("*.png"):
        try:
            imagem_antiga.unlink()
        except Exception:
            pass
    for doc_antigo in Path(PASTA_DOCUMENTOS_NF).glob("*"):
        try:
            if doc_antigo.is_file():
                doc_antigo.unlink()
        except Exception:
            pass

    todos_balancos, todas_categorias, todas_movimentacoes, todas_cobrancas = [], [], [], []
    for i, pdf in enumerate(pdfs):
        if progress_callback:
            pct = int((i + 1) / len(pdfs) * 60)
            progress_callback(pct, f'Processando {pdf.name}... ({i + 1}/{len(pdfs)})')
        balanco, categorias, movimentacoes, cobrancas = processar_pdf(pdf)
        if balanco: todos_balancos.append(balanco)
        todas_categorias.extend(categorias)
        todas_movimentacoes.extend(movimentacoes)
        todas_cobrancas.extend(cobrancas)

    if progress_callback:
        progress_callback(62, 'Consolidando dataframes...')
    df_balanco = criar_dataframe(todos_balancos, COLUNAS_BALANCO).drop_duplicates(subset=["Hash"]).sort_values(by=["Mes_Ano", "Arquivo_Origem"]) if todos_balancos else pd.DataFrame(columns=COLUNAS_BALANCO)
    df_cat = criar_dataframe(todas_categorias, COLUNAS_CATEGORIAS).drop_duplicates(subset=["Hash"]).sort_values(by=["Mes_Ano", "Grupo", "Categoria"]) if todas_categorias else pd.DataFrame(columns=COLUNAS_CATEGORIAS)
    df_mov = criar_dataframe(todas_movimentacoes, COLUNAS_MOVIMENTACOES).drop_duplicates(subset=["Hash"]).sort_values(by=["Mes_Ano", "Data", "Conta", "Tipo_Movimento"]) if todas_movimentacoes else pd.DataFrame(columns=COLUNAS_MOVIMENTACOES)
    if progress_callback:
        progress_callback(68, 'Recalculando scores de divergência...')
    df_mov = recalcular_scores_divergencia(df_mov)
    df_cobrancas = criar_dataframe(todas_cobrancas, COLUNAS_COBRANCAS).drop_duplicates(subset=["Hash"]).sort_values(by=["Mes_Ano", "Data_Liquidacao", "Bloco", "Unidade"]) if todas_cobrancas else pd.DataFrame(columns=COLUNAS_COBRANCAS)
    if progress_callback:
        progress_callback(74, 'Detectando fraudes...')
    df_fraudes = detectar_fraudes(df_mov)
    df_validacao = criar_validacao(df_balanco, df_mov)
    if progress_callback:
        progress_callback(80, 'Exportando Excel...')
    exportar_excel(df_balanco, df_cat, df_mov, df_cobrancas, df_fraudes, df_validacao)
    if progress_callback:
        progress_callback(88, 'Exportando CSVs...')
    exportar_csvs(df_balanco, df_cat, df_mov, df_cobrancas)
    if progress_callback:
        progress_callback(95, 'Criando dashboards HTML...')
    criar_dashboards(df_balanco, df_cat, df_mov)
    if progress_callback:
        progress_callback(100, 'Auditoria concluída!')

if __name__ == "__main__":
    main()
