
# -*- coding: utf-8 -*-
from __future__ import annotations
import os, re, json, html, time, shutil, hashlib, unicodedata
from pathlib import Path
from difflib import SequenceMatcher
from urllib.parse import urljoin, urlparse
import pandas as pd
import fitz
import pytesseract
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
try:
    import streamlit as st
    import plotly.express as px
    import plotly.graph_objects as go
except Exception:
    st=None; px=None; go=None

# ============================================================#Image


# ============================================================
# CONFIGURAÇÕES
# ============================================================

PASTA_PDFS = "./ArquivosPDF"
PASTA_IMAGENS = "./notas_fiscais"

ARQUIVO_EXCEL = "auditoria_condominio.xlsx"

ARQUIVO_CSV_BALANCO = "auditoria_balanco_mensal.csv"
ARQUIVO_CSV_CATEGORIAS = "auditoria_categorias_mensais.csv"
ARQUIVO_CSV_MOVIMENTACOES = "auditoria_movimentacoes.csv"
ARQUIVO_CSV_COBRANCAS = "auditoria_composicao_cobrancas.csv"

os.makedirs(PASTA_IMAGENS, exist_ok=True)


# ============================================================
# OCR - WINDOWS
# ============================================================

# Se estiver no Windows e precisar:
#
# pytesseract.pytesseract.tesseract_cmd = (
#     r"C:\Program Files\Tesseract-OCR\tesseract.exe"
# )


# ============================================================
# REGEX
# ============================================================

REGEX_DATA = re.compile(r"\d{2}/\d{2}/\d{4}")

REGEX_VALOR_TXT = r"\(?-?\d{1,3}(?:\.\d{3})*,\d{2}\)?"

REGEX_VALOR = re.compile(REGEX_VALOR_TXT)

REGEX_CNPJ = re.compile(
    r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}"
)

REGEX_PERIODO = re.compile(
    r"(?:Período|Periodo|De)\s+(\d{2}/\d{2}/\d{4})\s+(?:a|até|ate)\s+(\d{2}/\d{2}/\d{4})",
    re.IGNORECASE
)


# ============================================================
# COLUNAS PADRÃO
# ============================================================

COLUNAS_BALANCO = [
    "Arquivo_Origem",
    "Condominio",
    "Periodo_Inicio",
    "Periodo_Fim",
    "Ano",
    "Mes",
    "Mes_Ano",
    "Saldo_Anterior",
    "Total_Receitas",
    "Total_Despesas",
    "Movimento_Liquido",
    "Saldo_Final",
    "Diferenca_Resultado",
    "Diferenca_Saldo",
    "Status_Validacao",
    "Hash",
]

COLUNAS_CATEGORIAS = [
    "Arquivo_Origem",
    "Condominio",
    "Periodo_Inicio",
    "Periodo_Fim",
    "Ano",
    "Mes",
    "Mes_Ano",
    "Grupo",
    "Categoria",
    "Valor",
    "Valor_Entrada",
    "Valor_Saida",
    "Valor_Resultante",
    "Origem_Extracao",
    "Hash",
]

COLUNAS_MOVIMENTACOES = [
    "Arquivo_Origem",
    "Condominio",
    "Periodo_Inicio",
    "Periodo_Fim",
    "Ano",
    "Mes",
    "Mes_Ano",
    "Data",
    "Conta",
    "Tipo_Movimento",
    "Categoria",
    "Fornecedor",
    "Descricao",
    "Valor_Real",
    "Valor_Assinado",
    "Valor_Entrada",
    "Valor_Saida",
    "Valor_Transferencia",
    "Saldo_Apos",
    "CNPJ",
    "PIX",
    "Score_Suspeita",
    "Motivo_Suspeita",
    "Link_NF",
    "Origem_Extracao",
    "Hash",
]

COLUNAS_COBRANCAS = [
    "Arquivo_Origem",
    "Condominio",
    "Periodo_Inicio",
    "Periodo_Fim",
    "Ano",
    "Mes",
    "Mes_Ano",
    "Unidade",
    "Bloco",
    "Codigo_Cobranca",
    "Descricao",
    "Categoria",
    "Vencimento",
    "Data_Credito",
    "Data_Liquidacao",
    "Valor",
    "Valor_Assinado",
    "Tipo_Movimento",
    "Origem_Extracao",
    "Hash",
]

COLUNAS_VALIDACAO = [
    "Arquivo_Origem",
    "Mes_Ano",
    "Receitas_Balancete",
    "Entradas_Movimentacao",
    "Diferenca_Entradas",
    "Despesas_Balancete",
    "Saidas_Movimentacao",
    "Diferenca_Saidas",
    "Movimento_Liquido_Balancete",
    "Saldo_Anterior",
    "Saldo_Final",
    "Status",
]


# ============================================================
# UTILITÁRIOS
# ============================================================

def limpar_texto(texto):
    if not texto:
        return ""

    texto = str(texto)
    texto = texto.replace("\n", " ")
    texto = texto.replace("\r", " ")
    texto = re.sub(r"\s+", " ", texto)
    texto = re.sub(r"_+", "", texto)

    return texto.strip()


def limpar_linha(texto):
    if not texto:
        return ""

    texto = str(texto)
    texto = texto.replace("\r", " ")
    texto = re.sub(r"\s+", " ", texto)

    return texto.strip()


def normalizar(texto):
    if texto is None:
        return ""

    texto = str(texto).lower()

    troca = {
        "á": "a",
        "à": "a",
        "ã": "a",
        "â": "a",
        "é": "e",
        "ê": "e",
        "í": "i",
        "ó": "o",
        "ô": "o",
        "õ": "o",
        "ú": "u",
        "ü": "u",
        "ç": "c",
    }

    for original, novo in troca.items():
        texto = texto.replace(original, novo)

    return texto


def valor_para_float(valor):
    if valor is None:
        return None

    valor = str(valor).strip()

    negativo = False

    if "(" in valor and ")" in valor:
        negativo = True

    if valor.startswith("-"):
        negativo = True

    valor = valor.replace("(", "")
    valor = valor.replace(")", "")
    valor = valor.replace("-", "")
    valor = valor.replace(".", "")
    valor = valor.replace(",", ".")

    try:
        valor_float = float(valor)
    except Exception:
        return None

    if negativo:
        valor_float *= -1

    return valor_float


def float_para_br(valor):
    try:
        if pd.isna(valor):
            return ""
    except Exception:
        pass

    try:
        return f"{float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return ""


def converter_data(data_txt):
    try:
        return pd.to_datetime(
            data_txt,
            format="%d/%m/%Y",
            errors="coerce"
        )
    except Exception:
        return pd.NaT


def gerar_hash(texto):
    return hashlib.md5(
        str(texto).encode("utf-8")
    ).hexdigest()


def criar_dataframe(registros, colunas):
    if registros:
        df = pd.DataFrame(registros)

        for col in colunas:
            if col not in df.columns:
                df[col] = None

        return df[colunas]

    return pd.DataFrame(columns=colunas)


# ============================================================
# FUNÇÕES CRÍTICAS PARA NÃO PEGAR PERCENTUAL COMO VALOR
# ============================================================

def extrair_valores_com_contexto(texto):
    """
    Retorna lista de dicionários:
    {
        "txt": "20.460,88",
        "valor": 20460.88,
        "eh_percentual": False
    }
    """

    resultados = []

    for m in REGEX_VALOR.finditer(texto):
        valor_txt = m.group(0)
        valor_float = valor_para_float(valor_txt)

        if valor_float is None:
            continue

        pos_fim = m.end()
        restante = texto[pos_fim:pos_fim + 5]

        eh_percentual = False

        if "%" in restante:
            eh_percentual = True

        resultados.append({
            "txt": valor_txt,
            "valor": valor_float,
            "eh_percentual": eh_percentual,
        })

    return resultados


def escolher_valor_monetario(texto):
    """
    Escolhe o valor monetário mais provável em um trecho.

    Regras:
    1) Ignora valores claramente percentuais quando seguidos de %
    2) Se sobrarem vários valores, pega o maior valor absoluto
       Exemplo:
       'Total de Receitas 20.460,88 100,00' -> 20.460,88
       'Total de Receitas 100,00% 20.460,88' -> 20.460,88
    """

    valores = extrair_valores_com_contexto(texto)

    if not valores:
        return None

    nao_percentuais = [
        item["valor"]
        for item in valores
        if not item["eh_percentual"]
    ]

    if not nao_percentuais:
        nao_percentuais = [
            item["valor"]
            for item in valores
        ]

    if not nao_percentuais:
        return None

    return max(
        nao_percentuais,
        key=lambda x: abs(x)
    )


def extrair_trecho_entre_rotulos(texto, rotulo_inicio, rotulos_fim):
    texto_norm = normalizar(texto)
    inicio_norm = normalizar(rotulo_inicio)

    pos_ini = texto_norm.find(inicio_norm)

    if pos_ini == -1:
        return ""

    pos_busca = pos_ini + len(inicio_norm)

    pos_fim_encontrado = None

    for rotulo_fim in rotulos_fim:
        fim_norm = normalizar(rotulo_fim)
        pos_fim = texto_norm.find(fim_norm, pos_busca)

        if pos_fim != -1:
            if pos_fim_encontrado is None or pos_fim < pos_fim_encontrado:
                pos_fim_encontrado = pos_fim

    if pos_fim_encontrado is None:
        return texto[pos_ini:]

    return texto[pos_ini:pos_fim_encontrado]


def buscar_valor_rotulo_por_trecho(texto, rotulo, proximos_rotulos):
    trecho = extrair_trecho_entre_rotulos(
        texto,
        rotulo,
        proximos_rotulos
    )

    if not trecho:
        return None

    return escolher_valor_monetario(trecho)


# ============================================================
# OCR
# ============================================================

def executar_ocr(pagina):
    try:
        pix = pagina.get_pixmap(
            matrix=fitz.Matrix(2, 2)
        )

        temp = "ocr_temp.png"

        pix.save(
            temp,
            output="png"
        )

        texto = pytesseract.image_to_string(
            Image.open(temp),
            lang="por"
        )

        os.remove(temp)

        return texto

    except Exception:
        return ""


# ============================================================
# EXTRAÇÃO DO PDF
# ============================================================

def extrair_paginas_pdf(pdf_path):
    doc = fitz.open(pdf_path)

    paginas = []

    for idx in range(len(doc)):
        pagina = doc[idx]

        texto = pagina.get_text("text")

        if len(texto.strip()) < 30:
            texto = executar_ocr(pagina)

        linhas = [
            limpar_linha(linha)
            for linha in texto.splitlines()
            if limpar_linha(linha)
        ]

        paginas.append({
            "pagina": idx + 1,
            "texto": texto,
            "texto_limpo": limpar_texto(texto),
            "linhas": linhas
        })

    return paginas


# ============================================================
# METADADOS
# ============================================================

def extrair_metadados(pdf_path, paginas):
    texto_total = "\n".join(
        pagina["texto"]
        for pagina in paginas[:10]
    )

    texto_limpo = limpar_texto(texto_total)

    periodo_ini = None
    periodo_fim = None
    ano = None
    mes = None
    mes_ano = None

    m_periodo = REGEX_PERIODO.search(texto_limpo)

    if m_periodo:
        periodo_ini = m_periodo.group(1)
        periodo_fim = m_periodo.group(2)

        dt = converter_data(periodo_fim)

        if pd.notna(dt):
            ano = int(dt.year)
            mes = int(dt.month)
            mes_ano = dt.strftime("%Y-%m")

    condominio = "NÃO IDENTIFICADO"

    m_condominio = re.search(
        r"Condomínio:\s*([A-ZÁÉÍÓÚÂÊÔÃÕÇ0-9\s]+)",
        texto_limpo,
        re.IGNORECASE
    )

    if m_condominio:
        condominio = limpar_texto(
            m_condominio.group(1)
        )

    if condominio == "NÃO IDENTIFICADO":
        m_condominio_alt = re.search(
            r"CONDOMÍNIO\s+RESIDENCIAL\s+([A-ZÁÉÍÓÚÂÊÔÃÕÇ0-9\s]+)",
            texto_limpo,
            re.IGNORECASE
        )

        if m_condominio_alt:
            condominio = "CONDOMÍNIO RESIDENCIAL " + limpar_texto(
                m_condominio_alt.group(1)
            )

    return {
        "Arquivo_Origem": pdf_path.name,
        "Condominio": condominio,
        "Periodo_Inicio": periodo_ini,
        "Periodo_Fim": periodo_fim,
        "Ano": ano,
        "Mes": mes,
        "Mes_Ano": mes_ano
    }


# ============================================================
# BALANÇO MENSAL
# ============================================================

def buscar_saldo_anterior(texto):
    trecho = extrair_trecho_entre_rotulos(
        texto,
        "Saldo Anterior",
        [
            "Receitas",
            "Total de Receitas",
            "Despesas Mensais",
        ]
    )

    if trecho:
        return escolher_valor_monetario(trecho)

    return None


def buscar_total_receitas(texto):
    trecho = extrair_trecho_entre_rotulos(
        texto,
        "Total de Receitas",
        [
            "Despesas Mensais",
            "Despesas",
            "Total de Mensais",
            "Resumo Financeiro",
            "SALDO ATUAL",
        ]
    )

    if trecho:
        return abs(escolher_valor_monetario(trecho))

    return None


def buscar_total_despesas(texto):
    trecho = extrair_trecho_entre_rotulos(
        texto,
        "Total de Despesas",
        [
            "Mov. Líquido",
            "Mov. Liquido",
            "Saldo em",
            "SALDO ATUAL",
            "Resumo Financeiro",
            "AHAVA",
        ]
    )

    if trecho:
        return abs(escolher_valor_monetario(trecho))

    return None


def buscar_movimento_liquido(texto):
    trecho = extrair_trecho_entre_rotulos(
        texto,
        "Mov. Líquido",
        [
            "Saldo em",
            "SALDO ATUAL",
            "Resumo Financeiro",
            "AHAVA",
        ]
    )

    if not trecho:
        trecho = extrair_trecho_entre_rotulos(
            texto,
            "Mov. Liquido",
            [
                "Saldo em",
                "SALDO ATUAL",
                "Resumo Financeiro",
                "AHAVA",
            ]
        )

    if trecho:
        return escolher_valor_monetario(trecho)

    return None


def buscar_saldo_final(texto, meta):
    periodo_fim = meta.get("Periodo_Fim")

    rotulos = []

    if periodo_fim:
        rotulos.append(f"Saldo em {periodo_fim}")

    rotulos.extend([
        "SALDO ATUAL (SALDO ANTERIOR + RECEITA - DESPESA)",
        "Saldo Atual",
        "Saldo Total",
    ])

    for rotulo in rotulos:
        trecho = extrair_trecho_entre_rotulos(
            texto,
            rotulo,
            [
                "Resumo dos Saldos",
                "AHAVA",
                "Conta Saldo",
                "Banco Sicoob",
                "Emitido em",
            ]
        )

        if trecho:
            valor = escolher_valor_monetario(trecho)

            if valor is not None:
                return valor

    return None


def extrair_balanco_mensal(pdf_path, paginas, meta):
    texto_primeiras = "\n".join(
        pagina["texto_limpo"]
        for pagina in paginas[:8]
    )

    saldo_anterior = buscar_saldo_anterior(
        texto_primeiras
    )

    total_receitas = buscar_total_receitas(
        texto_primeiras
    )

    total_despesas = buscar_total_despesas(
        texto_primeiras
    )

    movimento_liquido = buscar_movimento_liquido(
        texto_primeiras
    )

    saldo_final = buscar_saldo_final(
        texto_primeiras,
        meta
    )

    if movimento_liquido is None and total_receitas is not None and total_despesas is not None:
        movimento_liquido = round(
            total_receitas - total_despesas,
            2
        )

    if saldo_final is None and saldo_anterior is not None and movimento_liquido is not None:
        saldo_final = round(
            saldo_anterior + movimento_liquido,
            2
        )

    diferenca_resultado = None
    diferenca_saldo = None

    if total_receitas is not None and total_despesas is not None and movimento_liquido is not None:
        diferenca_resultado = round(
            movimento_liquido - (total_receitas - total_despesas),
            2
        )

    if saldo_anterior is not None and movimento_liquido is not None and saldo_final is not None:
        diferenca_saldo = round(
            saldo_final - (saldo_anterior + movimento_liquido),
            2
        )

    status_validacao = "OK"

    if diferenca_resultado not in [None, 0]:
        status_validacao = "DIVERGÊNCIA_RESULTADO"

    if diferenca_saldo not in [None, 0]:
        status_validacao = "DIVERGÊNCIA_SALDO"

    return {
        **meta,
        "Saldo_Anterior": saldo_anterior,
        "Total_Receitas": total_receitas,
        "Total_Despesas": total_despesas,
        "Movimento_Liquido": movimento_liquido,
        "Saldo_Final": saldo_final,
        "Diferenca_Resultado": diferenca_resultado,
        "Diferenca_Saldo": diferenca_saldo,
        "Status_Validacao": status_validacao,
        "Hash": gerar_hash(
            f"{pdf_path.name}_{meta.get('Mes_Ano')}_{saldo_anterior}_{total_receitas}_{total_despesas}_{saldo_final}"
        )
    }


# ============================================================
# CATEGORIAS MENSAIS
# ============================================================

RECEITAS_ROTULOS = [
    "Cotas do Mês",
    "Juros",
    "Multas",
    "Tarifa bancária",
    "Água",
    "Atualização Monetária",
    "Honorário de Cobrança",
    "Pagamentos a Maior",
    "Fundo de Obras",
    "Taxa fixa Copasa",
    "Outras Entradas",
]

DESPESAS_ROTULOS = [
    "Energia elétrica",
    "Água e Esgoto",
    "Seguro obrigatório",
    "Limpeza e conservação",
    "Limpeza Caixas de gordura - Água",
    "Câmeras e Interfones",
    "Manutenção Geral",
    "Locação de Equipamentos",
    "Dedetização - Controle de Pragas",
    "Combustível",
    "Conservadora",
    "Administradora de Condomínio",
    "Síndico Profissional",
    "Assessoria e Serviços Profissionais",
    "Chaveiro",
    "Taxa de Administração de Água e-ou Gás",
    "INSS Notas",
    "Honorários de cobrança",
    "Despesa com Processo Judicial",
    "Tarifa bancária",
    "Reformas em geral",
]


def extrair_secao(texto, inicio, fim=None):
    texto_norm = normalizar(texto)
    inicio_norm = normalizar(inicio)

    pos_ini = texto_norm.find(inicio_norm)

    if pos_ini == -1:
        return ""

    if fim:
        fim_norm = normalizar(fim)

        pos_fim = texto_norm.find(
            fim_norm,
            pos_ini + len(inicio_norm)
        )

        if pos_fim != -1:
            return texto[pos_ini:pos_fim]

    return texto[pos_ini:]


def extrair_categorias_mensais(pdf_path, paginas, meta):
    registros = []

    texto = "\n".join(
        pagina["texto_limpo"]
        for pagina in paginas[:8]
    )

    secao_receitas = extrair_secao(
        texto,
        "Receitas",
        "Despesas Mensais"
    )

    secao_despesas = extrair_secao(
        texto,
        "Despesas Mensais",
        "Resumo Financeiro"
    )

    if not secao_despesas:
        secao_despesas = extrair_secao(
            texto,
            "Despesas Mensais"
        )

    for idx, rotulo in enumerate(RECEITAS_ROTULOS):
        proximos = RECEITAS_ROTULOS[idx + 1:] + [
            "Total de Receitas",
            "Despesas Mensais",
        ]

        valor = buscar_valor_rotulo_por_trecho(
            secao_receitas,
            rotulo,
            proximos
        )

        if valor is not None:
            registros.append({
                **meta,
                "Grupo": "Entrada",
                "Categoria": rotulo,
                "Valor": abs(valor),
                "Valor_Entrada": abs(valor),
                "Valor_Saida": 0,
                "Valor_Resultante": valor,
                "Origem_Extracao": "Balancete/Categorias",
                "Hash": gerar_hash(
                    f"{pdf_path.name}_{meta.get('Mes_Ano')}_Entrada_{rotulo}_{valor}"
                )
            })

    for idx, rotulo in enumerate(DESPESAS_ROTULOS):
        proximos = DESPESAS_ROTULOS[idx + 1:] + [
            "Total de Mensais",
            "Total de Manutenção",
            "Total de Despesas Gerais",
            "Total de Prestação de serviços",
            "Total de Impostos e Taxas",
            "Total de Jurídico ou Cobrança",
            "Total de Despesas Bancárias",
            "Total de Despesas Obras e melhorias",
            "Total de Despesas",
            "Mov. Líquido",
            "Mov. Liquido",
        ]

        valor = buscar_valor_rotulo_por_trecho(
            secao_despesas,
            rotulo,
            proximos
        )

        if valor is not None:
            registros.append({
                **meta,
                "Grupo": "Saída",
                "Categoria": rotulo,
                "Valor": abs(valor),
                "Valor_Entrada": 0,
                "Valor_Saida": abs(valor),
                "Valor_Resultante": -abs(valor),
                "Origem_Extracao": "Balancete/Categorias",
                "Hash": gerar_hash(
                    f"{pdf_path.name}_{meta.get('Mes_Ano')}_Saida_{rotulo}_{valor}"
                )
            })

    return registros


# ============================================================
# CLASSIFICAÇÃO
# ============================================================

def classificar_categoria(descricao):
    t = normalizar(descricao)

    if "cemig" in t or "energia" in t:
        return "Energia"

    if "copasa" in t or "agua" in t or "esgoto" in t or "taxa fixa copasa" in t:
        return "Água"

    if "cota" in t:
        return "Cotas do Mês"

    if "fundo de obras" in t:
        return "Fundo de Obras"

    if "fundo de reserva" in t:
        return "Fundo de Reserva"

    if "tarifa bancaria" in t:
        return "Tarifa bancária"

    if "juros" in t:
        return "Juros"

    if "multa" in t:
        return "Multas"

    if "honorario" in t or "cobranca" in t:
        return "Honorários/Cobrança"

    if "receita federal" in t or "inss" in t or "darf" in t or "ministerio da fazenda" in t:
        return "Impostos"

    if (
        "manutencao" in t
        or "reparo" in t
        or "limpeza" in t
        or "material" in t
        or "dedetizacao" in t
        or "interfone" in t
        or "fechadura" in t
        or "obra" in t
        or "reforma" in t
        or "cacamba" in t
        or "locacamba" in t
    ):
        return "Manutenção/Obras"

    if "seguro" in t or "tokio" in t or "allianz" in t:
        return "Seguro"

    if "sindico" in t or "administradora" in t or "ahava" in t or "conservadora" in t:
        return "Prestação de Serviços"

    if "tribunal" in t or "processo judicial" in t:
        return "Jurídico"

    if "transferencia" in t or "transf." in t:
        return "Transferência"

    return "Outros"


def extrair_fornecedor(descricao):
    texto = descricao.upper()

    if " - " in descricao:
        return descricao.split(" - ")[0].strip()[:100]

    fornecedores = [
        "CEMIG",
        "COPASA",
        "RECEITA FEDERAL",
        "TOKIO MARINE",
        "ALLIANZ",
        "CONSERVADORA ATUANTE",
        "AHAVA",
        "TODAH",
        "MATOSO",
        "MGPRAG",
        "LOCAÇAMBA",
        "LOCACAMBA",
        "TRIBUNAL DE JUSTIÇA",
        "TRIBUNAL DE JUSTICA",
        "CHARLES",
        "PAULO HENRIQUE",
        "JOSE LUCIO",
        "VINICIUS SANTANA",
        "DELMA",
    ]

    for fornecedor in fornecedores:
        if fornecedor in texto:
            return fornecedor

    return ""


def extrair_cnpj(texto):
    m = REGEX_CNPJ.search(texto)

    if m:
        return m.group()

    return ""


def detectar_pix(texto):
    t = normalizar(texto)

    termos = [
        "pix",
        "chave pix",
        "transferencia",
        "ted",
        "cpf"
    ]

    return any(
        termo in t
        for termo in termos
    )


# ============================================================
# NOTAS
# ============================================================

def extrair_notas(pdf_path):
    print(f"\n📸 Extraindo comprovantes/notas: {pdf_path.name}")

    doc = fitz.open(pdf_path)

    mapa = {}

    for pagina_num in range(len(doc)):
        pagina = doc[pagina_num]

        texto = pagina.get_text("text")

        if len(texto.strip()) < 30:
            texto = executar_ocr(pagina)

        texto_lower = normalizar(texto)

        termos_comprovante = [
            "pix",
            "comprovante",
            "darf",
            "boleto",
            "pago a",
            "receita federal",
            "transferencia",
            "pagamento",
            "valor pago",
            "valor emitido",
            "valor total",
        ]

        if not any(termo in texto_lower for termo in termos_comprovante):
            continue

        valores = REGEX_VALOR.findall(texto)

        if not valores:
            continue

        valor = valor_para_float(
            valores[-1]
        )

        if valor is None:
            continue

        try:
            pix = pagina.get_pixmap(
                matrix=fitz.Matrix(2, 2)
            )

            valor_limpo = (
                str(abs(valor))
                .replace(".", "_")
                .replace("-", "")
            )

            nome = (
                f"NF_pag_{pagina_num + 1:03d}"
                f"_valor_{valor_limpo}.png"
            )

            caminho = os.path.join(
                PASTA_IMAGENS,
                nome
            )

            pix.save(
                caminho,
                output="png"
            )

            mapa[
                round(abs(valor), 2)
            ] = caminho

        except Exception:
            continue

    return mapa


# ============================================================
# MOVIMENTAÇÃO ANALÍTICA
# ============================================================

def limpar_linha_movimentacao(linha):
    linha = str(linha)

    linha = linha.replace("**", " ")
    linha = re.sub(r"AHAVA GEST.*", "", linha, flags=re.IGNORECASE)
    linha = re.sub(r"Rua Professor.*", "", linha, flags=re.IGNORECASE)
    linha = re.sub(r"Emitido em.*", "", linha, flags=re.IGNORECASE)
    linha = re.sub(r"\[.*?\]", "", linha)
    linha = re.sub(r"https?://\S+", "", linha)
    linha = re.sub(r"\s+", " ", linha)

    return linha.strip()


def identificar_conta_pagina(texto):
    m = re.search(
        r'Movimentação Analítica\s+"([^"]+)"',
        texto,
        re.IGNORECASE
    )

    if m:
        return m.group(1).strip()

    m = re.search(
        r'Movimentacao Analitica\s+"([^"]+)"',
        texto,
        re.IGNORECASE
    )

    if m:
        return m.group(1).strip()

    return ""


def tipo_movimento_por_valor(valor, descricao):
    desc_norm = normalizar(descricao)

    if "transf." in desc_norm or "transferencia" in desc_norm or "para a conta" in desc_norm:
        return "Transferência"

    if valor > 0:
        return "Entrada"

    if valor < 0:
        return "Saída"

    return "Neutro"


def score_suspeita_movimentacao(valor, descricao, link_nf):
    valor_abs = abs(valor)
    desc_norm = normalizar(descricao)

    score = 0
    motivos = []

    if valor >= 0:
        return 0, ""

    if valor_abs > 5000:
        score += 40
        motivos.append("Saída acima de R$ 5.000")

    if valor_abs > 500 and not link_nf:
        score += 20
        motivos.append("Saída acima de R$ 500 sem NF vinculada automaticamente")

    if detectar_pix(descricao):
        score += 15
        motivos.append("PIX/TED/CPF identificado")

    if "sem apresentacao" in desc_norm:
        score += 50
        motivos.append("Sem apresentação de NF")

    return score, " | ".join(motivos)


def limpar_texto_movimentacao_pagina(texto):
    texto = texto.replace("\n", " ")
    texto = texto.replace("\r", " ")

    texto = re.sub(r"https?://\S+", " ", texto)
    texto = re.sub(r"\[.*?\]", " ", texto)
    texto = re.sub(r"AHAVA GESTÃO CONDOMINIAL.*", " ", texto, flags=re.IGNORECASE)
    texto = re.sub(r"Rua Professor.*", " ", texto, flags=re.IGNORECASE)
    texto = re.sub(r"Contatos:.*", " ", texto, flags=re.IGNORECASE)
    texto = re.sub(r"CRA MG:.*", " ", texto, flags=re.IGNORECASE)
    texto = re.sub(r"\s+", " ", texto)

    return texto.strip()


def parse_movimentacoes_texto(texto):
    registros = []

    texto = limpar_texto_movimentacao_pagina(texto)

    padrao = re.compile(
        r"(\d{2}/\d{2}/\d{4})\s+"
        r"(.+?)\s+"
        r"(" + REGEX_VALOR_TXT + r")\s+"
        r"(" + REGEX_VALOR_TXT + r")"
        r"(?=\s+\d{2}/\d{2}/\d{4}|\s+AHAVA|\s+\d+\s+de\s+\d+|\s*$)",
        re.IGNORECASE
    )

    for m in padrao.finditer(texto):
        data = m.group(1)
        descricao = limpar_linha_movimentacao(m.group(2))
        valor_txt = m.group(3)
        saldo_txt = m.group(4)

        if not descricao:
            continue

        desc_norm = normalizar(descricao)

        ignorar_prefixos = [
            "data descricao valor saldo",
            "periodo em",
            "movimentacao analitica",
        ]

        if any(desc_norm.strip().startswith(item) for item in ignorar_prefixos):
            continue

        valor = valor_para_float(valor_txt)
        saldo = valor_para_float(saldo_txt)

        if valor is None:
            continue

        registros.append({
            "Data": data,
            "Descricao": descricao,
            "Valor_Movimento": valor,
            "Saldo_Apos": saldo
        })

    return registros


def extrair_movimentacoes_analiticas(pdf_path, paginas, meta, mapa_notas):
    registros = []

    for pagina in paginas:
        texto_pagina = pagina["texto"]
        texto_norm = normalizar(texto_pagina)

        if "movimentacao analitica" not in texto_norm:
            continue

        conta = identificar_conta_pagina(
            texto_pagina
        )

        movimentos = parse_movimentacoes_texto(
            texto_pagina
        )

        for mov in movimentos:
            data = mov["Data"]
            data_dt = converter_data(data)

            if pd.notna(data_dt):
                ano = int(data_dt.year)
                mes = int(data_dt.month)
                mes_ano = data_dt.strftime("%Y-%m")
            else:
                ano = meta.get("Ano")
                mes = meta.get("Mes")
                mes_ano = meta.get("Mes_Ano")

            descricao = mov["Descricao"]
            valor = mov["Valor_Movimento"]
            valor_abs = abs(valor)

            tipo = tipo_movimento_por_valor(
                valor,
                descricao
            )

            categoria = classificar_categoria(
                descricao
            )

            fornecedor = extrair_fornecedor(
                descricao
            )

            link_nf = mapa_notas.get(
                round(valor_abs, 2),
                ""
            )

            score, motivos = score_suspeita_movimentacao(
                valor,
                descricao,
                link_nf
            )

            registros.append({
                **meta,
                "Data": data,
                "Ano": ano,
                "Mes": mes,
                "Mes_Ano": mes_ano,
                "Conta": conta,
                "Tipo_Movimento": tipo,
                "Categoria": categoria,
                "Fornecedor": fornecedor,
                "Descricao": descricao,
                "Valor_Real": valor_abs,
                "Valor_Assinado": valor,
                "Valor_Entrada": valor_abs if tipo == "Entrada" else 0,
                "Valor_Saida": valor_abs if tipo == "Saída" else 0,
                "Valor_Transferencia": valor_abs if tipo == "Transferência" else 0,
                "Saldo_Apos": mov["Saldo_Apos"],
                "CNPJ": extrair_cnpj(descricao),
                "PIX": detectar_pix(descricao),
                "Score_Suspeita": score,
                "Motivo_Suspeita": motivos,
                "Link_NF": link_nf,
                "Origem_Extracao": "Movimentação Analítica",
                "Hash": gerar_hash(
                    f"{pdf_path.name}_{conta}_{data}_{descricao}_{valor}_{mov['Saldo_Apos']}"
                )
            })

    return registros


# ============================================================
# COMPOSIÇÃO DAS COBRANÇAS
# ============================================================

def extrair_composicao_cobrancas(pdf_path, paginas, meta):
    registros = []

    unidade_atual = ""
    bloco_atual = ""

    for pagina in paginas:
        texto_pagina = pagina["texto"]

        if "Composição das cobranças por crédito" not in texto_pagina:
            continue

        for linha in pagina["linhas"]:
            linha = limpar_linha_movimentacao(linha)

            m_unidade = re.match(
                r"^(\d{3})\s+(\d{2})\s+Cobrança",
                linha,
                re.IGNORECASE
            )

            if m_unidade:
                unidade_atual = m_unidade.group(1)
                bloco_atual = m_unidade.group(2)
                continue

            padrao_item = re.compile(
                r"^(\d{5,6})\s+(.+?)\s+"
                r"(\d{2}/\d{2}/\d{4})\s+"
                r"(\d{2}/\d{2}/\d{4})\s+"
                r"(\d{2}/\d{2}/\d{4})\s+"
                r"(" + REGEX_VALOR_TXT + r")\s*$"
            )

            m = padrao_item.match(linha)

            if not m:
                continue

            codigo_cobranca = m.group(1)
            descricao = limpar_linha(m.group(2))
            vencimento = m.group(3)
            data_credito = m.group(4)
            data_liquidacao = m.group(5)
            valor = valor_para_float(m.group(6))

            if valor is None:
                continue

            categoria = classificar_categoria(
                descricao
            )

            data_dt = converter_data(data_liquidacao)

            if pd.notna(data_dt):
                ano = int(data_dt.year)
                mes = int(data_dt.month)
                mes_ano = data_dt.strftime("%Y-%m")
            else:
                ano = meta.get("Ano")
                mes = meta.get("Mes")
                mes_ano = meta.get("Mes_Ano")

            registros.append({
                **meta,
                "Ano": ano,
                "Mes": mes,
                "Mes_Ano": mes_ano,
                "Unidade": unidade_atual,
                "Bloco": bloco_atual,
                "Codigo_Cobranca": codigo_cobranca,
                "Descricao": descricao,
                "Categoria": categoria,
                "Vencimento": vencimento,
                "Data_Credito": data_credito,
                "Data_Liquidacao": data_liquidacao,
                "Valor": abs(valor),
                "Valor_Assinado": valor,
                "Tipo_Movimento": "Entrada" if valor >= 0 else "Saída",
                "Origem_Extracao": "Composição das Cobranças",
                "Hash": gerar_hash(
                    f"{pdf_path.name}_{unidade_atual}_{bloco_atual}_{codigo_cobranca}_{descricao}_{vencimento}_{data_credito}_{data_liquidacao}_{valor}"
                )
            })

    return registros


# ============================================================
# VALIDAÇÃO
# ============================================================

def criar_validacao(df_balanco, df_mov):
    registros = []

    if df_balanco.empty:
        return pd.DataFrame(columns=COLUNAS_VALIDACAO)

    if df_mov.empty or "Arquivo_Origem" not in df_mov.columns:
        for _, row in df_balanco.iterrows():
            registros.append({
                "Arquivo_Origem": row.get("Arquivo_Origem"),
                "Mes_Ano": row.get("Mes_Ano"),
                "Receitas_Balancete": row.get("Total_Receitas"),
                "Entradas_Movimentacao": 0,
                "Diferenca_Entradas": None,
                "Despesas_Balancete": row.get("Total_Despesas"),
                "Saidas_Movimentacao": 0,
                "Diferenca_Saidas": None,
                "Movimento_Liquido_Balancete": row.get("Movimento_Liquido"),
                "Saldo_Anterior": row.get("Saldo_Anterior"),
                "Saldo_Final": row.get("Saldo_Final"),
                "Status": "SEM_MOVIMENTAÇÃO_ANALÍTICA_EXTRAÍDA"
            })

        return pd.DataFrame(
            registros,
            columns=COLUNAS_VALIDACAO
        )

    for _, row in df_balanco.iterrows():
        arquivo = row.get("Arquivo_Origem")
        mes_ano = row.get("Mes_Ano")

        mov_mes = df_mov[
            (df_mov["Arquivo_Origem"] == arquivo)
            & (df_mov["Mes_Ano"] == mes_ano)
        ].copy()

        entradas_mov = mov_mes["Valor_Entrada"].sum() if not mov_mes.empty else 0
        saidas_mov = mov_mes["Valor_Saida"].sum() if not mov_mes.empty else 0

        total_receitas = row.get("Total_Receitas", 0)
        total_despesas = row.get("Total_Despesas", 0)

        diff_entradas = None
        diff_saidas = None

        if pd.notna(total_receitas):
            diff_entradas = round(
                entradas_mov - total_receitas,
                2
            )

        if pd.notna(total_despesas):
            diff_saidas = round(
                saidas_mov - total_despesas,
                2
            )

        status = "OK"

        if diff_entradas not in [None, 0] or diff_saidas not in [None, 0]:
            status = "VERIFICAR_DIFERENÇA_MOVIMENTAÇÃO_X_BALANCETE"

        registros.append({
            "Arquivo_Origem": arquivo,
            "Mes_Ano": mes_ano,
            "Receitas_Balancete": total_receitas,
            "Entradas_Movimentacao": entradas_mov,
            "Diferenca_Entradas": diff_entradas,
            "Despesas_Balancete": total_despesas,
            "Saidas_Movimentacao": saidas_mov,
            "Diferenca_Saidas": diff_saidas,
            "Movimento_Liquido_Balancete": row.get("Movimento_Liquido"),
            "Saldo_Anterior": row.get("Saldo_Anterior"),
            "Saldo_Final": row.get("Saldo_Final"),
            "Status": status
        })

    return pd.DataFrame(
        registros,
        columns=COLUNAS_VALIDACAO
    )


# ============================================================
# FRAUDES
# ============================================================

def detectar_fraudes(df_mov):
    if df_mov.empty:
        return pd.DataFrame(columns=COLUNAS_MOVIMENTACOES)

    if "Tipo_Movimento" not in df_mov.columns or "Score_Suspeita" not in df_mov.columns:
        return pd.DataFrame(columns=COLUNAS_MOVIMENTACOES)

    return df_mov[
        (df_mov["Tipo_Movimento"] == "Saída")
        & (df_mov["Score_Suspeita"] >= 40)
    ].copy()


# ============================================================
# DASHBOARDS
# ============================================================

def criar_dashboards(df_balanco, df_cat, df_mov):
    if not df_balanco.empty:
        bal = df_balanco.copy()
        bal = bal.sort_values("Mes_Ano")

        bal_melt = bal.melt(
            id_vars=["Mes_Ano"],
            value_vars=[
                "Total_Receitas",
                "Total_Despesas",
                "Movimento_Liquido"
            ],
            var_name="Indicador",
            value_name="Valor"
        )

        bal_melt["Indicador"] = bal_melt["Indicador"].replace({
            "Total_Receitas": "Entradas do mês",
            "Total_Despesas": "Saídas do mês",
            "Movimento_Liquido": "Balanço do mês"
        })

        fig = px.bar(
            bal_melt,
            x="Mes_Ano",
            y="Valor",
            color="Indicador",
            barmode="group",
            title="Entradas, Saídas e Balanço do Mês"
        )

        fig.write_html(
            "dashboard_balanco_mensal.html"
        )

        fig_saldo = go.Figure()

        fig_saldo.add_trace(
            go.Scatter(
                x=bal["Mes_Ano"],
                y=bal["Saldo_Anterior"],
                mode="lines+markers",
                name="Saldo anterior"
            )
        )

        fig_saldo.add_trace(
            go.Scatter(
                x=bal["Mes_Ano"],
                y=bal["Saldo_Final"],
                mode="lines+markers",
                name="Saldo final"
            )
        )

        fig_saldo.update_layout(
            title="Saldo Anterior x Saldo Final",
            xaxis_title="Mês",
            yaxis_title="Valor"
        )

        fig_saldo.write_html(
            "dashboard_saldos.html"
        )

    if not df_cat.empty:
        cat = df_cat.copy()

        fig_cat = px.bar(
            cat,
            x="Categoria",
            y="Valor",
            color="Grupo",
            barmode="group",
            title="Categorias de Entradas e Saídas"
        )

        fig_cat.write_html(
            "dashboard_categorias_entrada_saida.html"
        )

        fig_cat_mes = px.bar(
            cat,
            x="Mes_Ano",
            y="Valor",
            color="Categoria",
            facet_col="Grupo",
            title="Categorias por Mês - Entradas e Saídas"
        )

        fig_cat_mes.write_html(
            "dashboard_categorias_por_mes.html"
        )

    if not df_mov.empty and "Tipo_Movimento" in df_mov.columns:
        mov = df_mov[
            df_mov["Tipo_Movimento"].isin(["Entrada", "Saída", "Transferência"])
        ].copy()

        if not mov.empty:
            mov_group = (
                mov
                .groupby(["Mes_Ano", "Conta", "Tipo_Movimento"])[
                    ["Valor_Entrada", "Valor_Saida", "Valor_Transferencia"]
                ]
                .sum()
                .reset_index()
            )

            def escolher_valor(row):
                if row["Tipo_Movimento"] == "Entrada":
                    return row["Valor_Entrada"]

                if row["Tipo_Movimento"] == "Saída":
                    return row["Valor_Saida"]

                return row["Valor_Transferencia"]

            mov_group["Valor"] = mov_group.apply(
                escolher_valor,
                axis=1
            )

            fig_contas = px.bar(
                mov_group,
                x="Mes_Ano",
                y="Valor",
                color="Tipo_Movimento",
                facet_col="Conta",
                title="Movimentação Analítica por Conta"
            )

            fig_contas.write_html(
                "dashboard_movimentacao_por_conta.html"
            )

            top_saida = (
                mov[mov["Tipo_Movimento"] == "Saída"]
                .groupby("Fornecedor")["Valor_Saida"]
                .sum()
                .sort_values(ascending=False)
                .head(15)
                .reset_index()
            )

            if not top_saida.empty:
                fig_top = px.bar(
                    top_saida,
                    x="Fornecedor",
                    y="Valor_Saida",
                    title="Top Fornecedores por Saída"
                )

                fig_top.write_html(
                    "dashboard_top_fornecedores_saida.html"
                )


# ============================================================
# EXPORTAÇÃO
# ============================================================

def exportar_excel(
    df_balanco,
    df_cat,
    df_mov,
    df_cobrancas,
    df_fraudes,
    df_validacao
):
    with pd.ExcelWriter(
        ARQUIVO_EXCEL,
        engine="openpyxl"
    ) as writer:

        df_balanco.to_excel(
            writer,
            sheet_name="Balanco_Mensal",
            index=False
        )

        df_cat.to_excel(
            writer,
            sheet_name="Categorias_Mensais",
            index=False
        )

        df_mov.to_excel(
            writer,
            sheet_name="Movimentacoes",
            index=False
        )

        if not df_mov.empty and "Tipo_Movimento" in df_mov.columns:
            entradas = df_mov[
                df_mov["Tipo_Movimento"] == "Entrada"
            ].copy()

            saidas = df_mov[
                df_mov["Tipo_Movimento"] == "Saída"
            ].copy()

            transferencias = df_mov[
                df_mov["Tipo_Movimento"] == "Transferência"
            ].copy()
        else:
            entradas = pd.DataFrame(columns=COLUNAS_MOVIMENTACOES)
            saidas = pd.DataFrame(columns=COLUNAS_MOVIMENTACOES)
            transferencias = pd.DataFrame(columns=COLUNAS_MOVIMENTACOES)

        entradas.to_excel(
            writer,
            sheet_name="Entradas_Analiticas",
            index=False
        )

        saidas.to_excel(
            writer,
            sheet_name="Saidas_Analiticas",
            index=False
        )

        transferencias.to_excel(
            writer,
            sheet_name="Transferencias",
            index=False
        )

        df_cobrancas.to_excel(
            writer,
            sheet_name="Composicao_Cobrancas",
            index=False
        )

        df_fraudes.to_excel(
            writer,
            sheet_name="Fraudes",
            index=False
        )

        df_validacao.to_excel(
            writer,
            sheet_name="Validacao",
            index=False
        )

    formatar_excel()


def formatar_excel():
    wb = load_workbook(
        ARQUIVO_EXCEL
    )

    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]

        for cell in ws[1]:
            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.fill = PatternFill(
                start_color="1F4E78",
                end_color="1F4E78",
                fill_type="solid"
            )

        ws.freeze_panes = "A2"

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                try:
                    if cell.value is not None:
                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )
                except Exception:
                    pass

            ws.column_dimensions[col_letter].width = min(
                max_length + 2,
                70
            )

    if "Movimentacoes" in wb.sheetnames:
        ws = wb["Movimentacoes"]

        colunas = {
            cell.value: idx
            for idx, cell in enumerate(ws[1], start=1)
        }

        if "Link_NF" in colunas:
            col_nf = colunas["Link_NF"]
            coluna_img = ws.max_column + 1

            ws.cell(
                row=1,
                column=coluna_img,
                value="Miniatura_NF"
            )

            ws.cell(
                row=1,
                column=coluna_img
            ).font = Font(
                bold=True,
                color="FFFFFF"
            )

            ws.cell(
                row=1,
                column=coluna_img
            ).fill = PatternFill(
                start_color="1F4E78",
                end_color="1F4E78",
                fill_type="solid"
            )

            for linha in range(2, ws.max_row + 1):
                caminho = ws.cell(
                    linha,
                    col_nf
                ).value

                if caminho and os.path.exists(caminho):
                    try:
                        img = XLImage(caminho)

                        img.width = 70
                        img.height = 70

                        celula_img = ws.cell(
                            row=linha,
                            column=coluna_img
                        ).coordinate

                        ws.add_image(
                            img,
                            celula_img
                        )

                        ws.row_dimensions[
                            linha
                        ].height = 60

                    except Exception:
                        pass

    wb.save(
        ARQUIVO_EXCEL
    )


def exportar_csvs(
    df_balanco,
    df_cat,
    df_mov,
    df_cobrancas
):
    df_balanco.to_csv(
        ARQUIVO_CSV_BALANCO,
        sep=";",
        index=False,
        encoding="utf-8-sig"
    )

    df_cat.to_csv(
        ARQUIVO_CSV_CATEGORIAS,
        sep=";",
        index=False,
        encoding="utf-8-sig"
    )

    df_mov.to_csv(
        ARQUIVO_CSV_MOVIMENTACOES,
        sep=";",
        index=False,
        encoding="utf-8-sig"
    )

    df_cobrancas.to_csv(
        ARQUIVO_CSV_COBRANCAS,
        sep=";",
        index=False,
        encoding="utf-8-sig"
    )


# ============================================================
# PROCESSAMENTO DE UM PDF
# ============================================================

def processar_pdf(pdf_path):
    print("\n" + "=" * 80)
    print(f"🚀 PROCESSANDO {pdf_path.name}")
    print("=" * 80)

    paginas = extrair_paginas_pdf(
        pdf_path
    )

    meta = extrair_metadados(
        pdf_path,
        paginas
    )

    mapa_notas = extrair_notas(
        pdf_path
    )

    balanco = extrair_balanco_mensal(
        pdf_path,
        paginas,
        meta
    )

    categorias = extrair_categorias_mensais(
        pdf_path,
        paginas,
        meta
    )

    movimentacoes = extrair_movimentacoes_analiticas(
        pdf_path,
        paginas,
        meta,
        mapa_notas
    )

    cobrancas = extrair_composicao_cobrancas(
        pdf_path,
        paginas,
        meta
    )

    print(f"✅ Balanço extraído: {1 if balanco else 0}")
    print(f"   Receitas: {float_para_br(balanco.get('Total_Receitas'))}")
    print(f"   Despesas: {float_para_br(balanco.get('Total_Despesas'))}")
    print(f"   Movimento líquido: {float_para_br(balanco.get('Movimento_Liquido'))}")
    print(f"✅ Categorias extraídas: {len(categorias)}")
    print(f"✅ Movimentações extraídas: {len(movimentacoes)}")
    print(f"✅ Cobranças extraídas: {len(cobrancas)}")

    return balanco, categorias, movimentacoes, cobrancas


# ============================================================
# MAIN
# ============================================================

def main():
    pdfs = list(
        Path(PASTA_PDFS).glob("*.pdf")
    )

    if not pdfs:
        print("❌ Nenhum PDF encontrado em ./ArquivosPDF")
        return

    todos_balancos = []
    todas_categorias = []
    todas_movimentacoes = []
    todas_cobrancas = []

    for pdf in pdfs:
        balanco, categorias, movimentacoes, cobrancas = processar_pdf(
            pdf
        )

        if balanco:
            todos_balancos.append(
                balanco
            )

        todas_categorias.extend(
            categorias
        )

        todas_movimentacoes.extend(
            movimentacoes
        )

        todas_cobrancas.extend(
            cobrancas
        )

    df_balanco = criar_dataframe(
        todos_balancos,
        COLUNAS_BALANCO
    )

    df_cat = criar_dataframe(
        todas_categorias,
        COLUNAS_CATEGORIAS
    )

    df_mov = criar_dataframe(
        todas_movimentacoes,
        COLUNAS_MOVIMENTACOES
    )

    df_cobrancas = criar_dataframe(
        todas_cobrancas,
        COLUNAS_COBRANCAS
    )

    if not df_balanco.empty:
        df_balanco = df_balanco.drop_duplicates(
            subset=["Hash"]
        )

    if not df_cat.empty:
        df_cat = df_cat.drop_duplicates(
            subset=["Hash"]
        )

    if not df_mov.empty:
        df_mov = df_mov.drop_duplicates(
            subset=["Hash"]
        )

    if not df_cobrancas.empty:
        df_cobrancas = df_cobrancas.drop_duplicates(
            subset=["Hash"]
        )

    if not df_balanco.empty:
        df_balanco = df_balanco.sort_values(
            by=["Mes_Ano", "Arquivo_Origem"]
        )

    if not df_cat.empty:
        df_cat = df_cat.sort_values(
            by=["Mes_Ano", "Grupo", "Categoria"]
        )

    if not df_mov.empty:
        df_mov = df_mov.sort_values(
            by=["Mes_Ano", "Data", "Conta", "Tipo_Movimento"]
        )

    if not df_cobrancas.empty:
        df_cobrancas = df_cobrancas.sort_values(
            by=["Mes_Ano", "Data_Liquidacao", "Bloco", "Unidade"]
        )

    df_fraudes = detectar_fraudes(
        df_mov
    )

    df_validacao = criar_validacao(
        df_balanco,
        df_mov
    )

    exportar_excel(
        df_balanco,
        df_cat,
        df_mov,
        df_cobrancas,
        df_fraudes,
        df_validacao
    )

    exportar_csvs(
        df_balanco,
        df_cat,
        df_mov,
        df_cobrancas
    )

    criar_dashboards(
        df_balanco,
        df_cat,
        df_mov
    )

    total_receitas = df_balanco["Total_Receitas"].sum() if not df_balanco.empty else 0
    total_despesas = df_balanco["Total_Despesas"].sum() if not df_balanco.empty else 0
    resultado = df_balanco["Movimento_Liquido"].sum() if not df_balanco.empty else 0

    print("\n" + "=" * 80)
    print("✅ AUDITORIA FINALIZADA")
    print("=" * 80)

    print(f"📊 Excel: {ARQUIVO_EXCEL}")
    print(f"📄 CSV Balanço: {ARQUIVO_CSV_BALANCO}")
    print(f"📄 CSV Categorias: {ARQUIVO_CSV_CATEGORIAS}")
    print(f"📄 CSV Movimentações: {ARQUIVO_CSV_MOVIMENTACOES}")
    print(f"📄 CSV Cobranças: {ARQUIVO_CSV_COBRANCAS}")

    print("\n📈 Dashboards gerados:")
    print("   - dashboard_balanco_mensal.html")
    print("   - dashboard_saldos.html")
    print("   - dashboard_categorias_entrada_saida.html")
    print("   - dashboard_categorias_por_mes.html")
    print("   - dashboard_movimentacao_por_conta.html")
    print("   - dashboard_top_fornecedores_saida.html")

    print("\nResumo contábil oficial extraído:")
    print(f"✅ Total de receitas: R$ {float_para_br(total_receitas)}")
    print(f"✅ Total de despesas: R$ {float_para_br(total_despesas)}")
    print(f"✅ Resultado líquido: R$ {float_para_br(resultado)}")

    print("\nResumo de registros:")
    print(f"✅ Balanços mensais: {len(df_balanco)}")
    print(f"✅ Categorias mensais: {len(df_cat)}")
    print(f"✅ Movimentações analíticas: {len(df_mov)}")
    print(f"✅ Composição de cobranças: {len(df_cobrancas)}")
    print(f"✅ Suspeitas/fraudes: {len(df_fraudes)}")
    print(f"✅ Validações: {len(df_validacao)}")

    if df_balanco.empty:
        print("\n⚠️ Nenhum balanço mensal foi extraído.")
    else:
        print("\n✅ Conferência rápida do balanço:")
        for _, row in df_balanco.iterrows():
            print(
                f"   {row['Mes_Ano']} | "
                f"Receitas: R$ {float_para_br(row['Total_Receitas'])} | "
                f"Despesas: R$ {float_para_br(row['Total_Despesas'])} | "
                f"Resultado: R$ {float_para_br(row['Movimento_Liquido'])}"
            )


# ============================================================
# EXECUÇÃO
# ============================================================

if __name__ == "__main__":
    main()

